"""
IO helpers — Excel reading and writing.
=========================================

Reads the input dataframe and unit mapping from the user's Excel file,
and writes the cleaned-and-calibrated v3 workbook that the v2 engine
will then consume.
"""

from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .constants import CONFIG_RANGES
from .auto_calibration import CalibrationReport


def _write_processed_workbook(orig_path: str, out_path: str,
                              cleaned_df: pd.DataFrame,
                              calib: CalibrationReport,
                              run_mode: str) -> None:
    """
    Build a new xlsx with:
      - Sheet1 = cleaned data
      - DIAGNOSTIC_CONFIG with extra columns for transparency
      - All other sheets copied as-is
    """
    src = load_workbook(orig_path, read_only=False, data_only=True)
    dst = Workbook()
    # Remove default sheet
    dst.remove(dst.active)

    data_sheet_name = None
    for sname in src.sheetnames:
        first = src[sname].cell(row=1, column=1).value
        if first and str(first).strip().upper() == "TIMESTAMP":
            data_sheet_name = sname
            break
    if data_sheet_name is None:
        data_sheet_name = "Sheet1" if "Sheet1" in src.sheetnames else \
                          src.sheetnames[0]

    # Write data sheet from cleaned_df
    ws = dst.create_sheet(data_sheet_name)
    ws.append(["TIMESTAMP"] + [c for c in cleaned_df.columns
                                if c != "TIMESTAMP"])
    cols_order = ["TIMESTAMP"] + [c for c in cleaned_df.columns
                                   if c != "TIMESTAMP"]
    for _, row in cleaned_df[cols_order].iterrows():
        out_row = []
        for v in row.values:
            if isinstance(v, float) and not np.isfinite(v):
                out_row.append(None)
            else:
                out_row.append(v)
        ws.append(out_row)

    # Copy other sheets, with DIAGNOSTIC_CONFIG getting the new columns
    for sname in src.sheetnames:
        if sname == data_sheet_name:
            continue
        if sname == "DIAGNOSTIC_CONFIG":
            ws_new = dst.create_sheet("DIAGNOSTIC_CONFIG")
            # Header
            header = ["Parameter", "Value", "Description",
                      "Manual_Value", "Auto_Value", "Mode_Used",
                      "Auto_Confidence"]
            ws_new.append(header)
            for c_idx, _ in enumerate(header, start=1):
                ws_new.cell(row=1, column=c_idx).font = Font(bold=True)

            # Build a map of decisions for fast lookup
            d_by_param = {d.parameter: d for d in calib.decisions}

            # Pull original Description column from src
            desc_map: Dict[str, str] = {}
            src_ws = src["DIAGNOSTIC_CONFIG"]
            for r in src_ws.iter_rows(min_row=2, values_only=True):
                if r and r[0] is not None:
                    desc_map[str(r[0]).strip()] = str(r[2]) if len(r) > 2 and r[2] else ""

            # Use the order from CONFIG_RANGES, then any extras
            written = set()
            for param in CONFIG_RANGES.keys():
                d = d_by_param.get(param)
                if d is None:
                    continue
                ws_new.append([
                    param, d.used_value, desc_map.get(param, ""),
                    d.manual_value,
                    "" if d.auto_value is None else d.auto_value,
                    d.mode, d.confidence
                ])
                written.add(param)

            # Any param in src but not in CONFIG_RANGES — preserve verbatim
            for r in src_ws.iter_rows(min_row=2, values_only=True):
                if not r or r[0] is None:
                    continue
                p = str(r[0]).strip()
                if p in written:
                    continue
                ws_new.append([p, r[1] if len(r) > 1 else None,
                               r[2] if len(r) > 2 else "",
                               r[1] if len(r) > 1 else None,
                               "", "MANUAL", "N/A"])

            # column widths
            for c_idx, w in enumerate([28, 14, 50, 14, 14, 12, 12], start=1):
                ws_new.column_dimensions[get_column_letter(c_idx)].width = w
        else:
            # Copy sheet as-is
            src_ws = src[sname]
            ws_new = dst.create_sheet(sname)
            for r in src_ws.iter_rows(values_only=True):
                ws_new.append(list(r))

    dst.save(out_path)


# ══════════════════════════════════════════════════════════════════════
# MAIN ORCHESTRATOR
# ══════════════════════════════════════════════════════════════════════

def _read_unit_mapping(path: str) -> Dict[str, str]:
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        if "UNIT_MAPPING" not in wb.sheetnames:
            return {}
        ws = wb["UNIT_MAPPING"]
        out: Dict[str, str] = {}
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r and r[0] and r[1]:
                out[str(r[0]).strip()] = str(r[1]).strip()
        return out
    except Exception:
        return {}


def _build_dataframe(path: str, sheet_info: Dict[str, Any]) -> pd.DataFrame:
    """Pull the data sheet into a pandas DataFrame, with TIMESTAMP first."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_info["data_sheet"]]
    header_row = sheet_info["header_row"]
    headers = [c for c in sheet_info["headers"] if c is not None]
    rows = []
    for r in ws.iter_rows(min_row=header_row + 1, values_only=True):
        rows.append(list(r[:len(headers)]))
    df = pd.DataFrame(rows, columns=headers)
    # Coerce TIMESTAMP
    df["TIMESTAMP"] = pd.to_datetime(df["TIMESTAMP"], errors="coerce")
    df = df.dropna(subset=["TIMESTAMP"]).reset_index(drop=True)
    return df
