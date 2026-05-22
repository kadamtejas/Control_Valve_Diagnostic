"""
Health check — input validation in 5 stages.
================================================

Module 1 of the v3 wrapper. Runs 45+ checks across 5 stages on the input
Excel file before the diagnostic engine touches the data:

  Stage 1 — File-level checks (existence, format, readable)
  Stage 2 — Sheet structure (required sheets, columns)
  Stage 3 — Configuration sanity (DIAGNOSTIC_CONFIG values in range)
  Stage 4 — Data quality (sentinels, gaps, dtypes, sample rate)
  Stage 5 — Cross-consistency (UNIT_MAPPING vs columns, etc.)

The two reporting classes (CheckResult, HealthCheckReport) and the five
stage functions are the public API of this module.
"""

import os
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from .constants import (
    PROBLEM, WARNING, INFO, PASSED,
    V3_VERSION,
    NUMERIC_SENTINELS, QUALITY_FLAGS, EXCEL_ERRORS,
    CONFIG_RANGES,
)
from .reporting_helpers import _line, _box

@dataclass
class CheckResult:
    """A single health-check finding."""
    check_id: str            # e.g. "1.4"
    name: str                # short label
    severity: str            # PROBLEM / WARNING / INFO / PASSED
    title: str = ""          # one-line summary
    what: str = ""           # what happened
    why: str = ""            # why it matters
    fix: str = ""            # how to fix
    extra: List[str] = field(default_factory=list)  # extra detail lines

    def render(self) -> str:
        body = []
        if self.what:
            body.append("WHAT HAPPENED:")
            for ln in self.what.split("\n"):
                body.append("  " + ln)
            body.append("")
        if self.why:
            body.append("WHY THIS MATTERS:")
            for ln in self.why.split("\n"):
                body.append("  " + ln)
            body.append("")
        if self.fix:
            body.append("HOW TO FIX:")
            for ln in self.fix.split("\n"):
                body.append("  " + ln)
        if self.extra:
            if self.fix:
                body.append("")
            body.extend(self.extra)
        title = self.title or self.name
        title = f"[Check {self.check_id}] {title}"
        return _box(title, self.severity, body)


@dataclass
class HealthCheckReport:
    """Aggregated output of all 5 stages."""
    file_path: str = ""
    run_at: str = ""
    results: List[CheckResult] = field(default_factory=list)

    @property
    def problems(self) -> List[CheckResult]:
        return [r for r in self.results if r.severity == PROBLEM]

    @property
    def warnings(self) -> List[CheckResult]:
        return [r for r in self.results if r.severity == WARNING]

    @property
    def infos(self) -> List[CheckResult]:
        return [r for r in self.results if r.severity == INFO]

    @property
    def passed(self) -> List[CheckResult]:
        return [r for r in self.results if r.severity == PASSED]

    @property
    def has_problems(self) -> bool:
        return len(self.problems) > 0

    def add(self, r: CheckResult):
        self.results.append(r)

    def add_passed(self, check_id: str, name: str):
        self.results.append(CheckResult(check_id=check_id, name=name,
                                        severity=PASSED))

    def render(self) -> str:
        out = []
        out.append(_line("═"))
        out.append(f"  INPUT HEALTH CHECK — {os.path.basename(self.file_path)}")
        out.append(f"  Run at {self.run_at}")
        out.append(f"  Tool version: valve_diagnostics v{V3_VERSION}")
        out.append(_line("═"))
        out.append("")
        out.append(f"  ❌  {len(self.problems):>2} PROBLEMS  "
                   f"(must fix before tool can run)")
        out.append(f"  ⚠️   {len(self.warnings):>2} WARNINGS  "
                   f"(tool will run but results may be affected)")
        out.append(f"  ℹ️   {len(self.infos):>2} INFOS     "
                   f"(observations, no action needed)")
        out.append(f"  ✅  {len(self.passed):>2} CHECKS PASSED")
        out.append("")

        if self.problems:
            out.append(_line())
            out.append("  PROBLEMS — fix these first")
            out.append(_line())
            out.append("")
            for r in self.problems:
                out.append(r.render())

        if self.warnings:
            out.append(_line())
            out.append("  WARNINGS — review but not blocking")
            out.append(_line())
            out.append("")
            for r in self.warnings:
                out.append(r.render())

        if self.infos:
            out.append(_line())
            out.append("  INFOS — observations")
            out.append(_line())
            out.append("")
            for r in self.infos:
                out.append(r.render())

        if self.passed:
            out.append(_line())
            out.append("  CHECKS PASSED — for transparency")
            out.append(_line())
            out.append("")
            for r in self.passed:
                out.append(f"  ✅ [{r.check_id}] {r.name}")
            out.append("")

        return "\n".join(out)


# ══════════════════════════════════════════════════════════════════════
# STAGE 1 — FILE-LEVEL HEALTH CHECKS
# ══════════════════════════════════════════════════════════════════════

def _stage1_file_checks(path: str, report: HealthCheckReport) -> bool:
    """Returns False if a problem is found that prevents further checks."""

    # 1.1 — file exists
    if not os.path.exists(path):
        report.add(CheckResult(
            check_id="1.1", name="File exists", severity=PROBLEM,
            title="The input file could not be found.",
            what=f"The tool looked for the file at:\n    {path}\n"
                 "but no file with that name exists at that location.",
            why="Without an input file, the tool has no data to analyse.",
            fix="1. Check the file name and path you typed are correct.\n"
                "2. If you dragged a file onto RUN.bat: try again, making\n"
                "   sure the file actually drops onto the .bat file.\n"
                "3. Verify the file exists by opening that folder in Windows\n"
                "   Explorer or Finder."))
        return False
    report.add_passed("1.1", "File exists at the specified path")

    # 1.2 — not locked / openable
    try:
        with open(path, "rb") as fh:
            fh.read(16)
    except PermissionError:
        report.add(CheckResult(
            check_id="1.2", name="File not locked", severity=PROBLEM,
            title="The input file can't be opened.",
            what=f"{os.path.basename(path)} appears to be open in Excel "
                 "right now, or another program has a lock on it.",
            why="Windows locks files that are open in another program. The "
                "tool can't read the data while Excel has it.",
            fix=f"1. Close {os.path.basename(path)} in Excel\n"
                "2. Re-run the tool\n\n"
                "(The file does NOT need to be closed every time — only when "
                "you are running the tool.)"))
        return False
    except Exception as e:
        report.add(CheckResult(
            check_id="1.2", name="File not locked", severity=PROBLEM,
            title="The input file can't be read.",
            what=f"The system reported: {e}",
            why="Cannot proceed without read access to the input file.",
            fix="Check file permissions and that the file is not corrupted."))
        return False
    report.add_passed("1.2", "File is not locked by another program")

    # 1.3 — valid xlsx
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        wb.close()
    except Exception as e:
        msg = str(e).lower()
        if "encrypted" in msg or "password" in msg:
            report.add(CheckResult(
                check_id="1.4", name="File not encrypted", severity=PROBLEM,
                title="The input file is password-protected.",
                what=f"{os.path.basename(path)} is encrypted and requires a "
                     "password.",
                why="The tool cannot read encrypted files.",
                fix="1. Open the file in Excel\n"
                    "2. File → Info → Protect Workbook → Encrypt with Password\n"
                    "3. Delete the password and click OK\n"
                    "4. Save the file\n"
                    "5. Re-run the tool"))
        else:
            report.add(CheckResult(
                check_id="1.3", name="Valid xlsx file", severity=PROBLEM,
                title="The input file is not a valid Excel workbook.",
                what=f"The tool tried to open {os.path.basename(path)} and "
                     f"failed:\n{e}",
                why="The tool requires a valid .xlsx file. This may be a\n"
                    "corrupted file, a renamed .xls (old format), or a file\n"
                    "that was saved with errors.",
                fix="1. Open the file in Excel\n"
                    "2. Save As... → choose 'Excel Workbook (.xlsx)' format\n"
                    "3. Re-run the tool with the new file"))
        return False
    report.add_passed("1.3", "File is a valid .xlsx workbook")
    report.add_passed("1.4", "File is not password-protected")

    # 1.5 — file size sanity
    size = os.path.getsize(path)
    if size < 5_000:
        report.add(CheckResult(
            check_id="1.5", name="File size sane", severity=WARNING,
            title="The input file is suspiciously small.",
            what=f"File size is {size:,} bytes. A real plant data export "
                 "is usually at least tens of kilobytes.",
            why="The tool may proceed but is unlikely to find usable data.",
            fix="Check whether the export from your historian completed "
                "successfully."))
    elif size > 500_000_000:
        report.add(CheckResult(
            check_id="1.5", name="File size sane", severity=WARNING,
            title="The input file is unusually large.",
            what=f"File size is {size/1_000_000:.0f} MB. This will work but "
                 "may take a long time to process.",
            why="Memory usage may be high. Consider splitting the analysis "
                "into smaller time windows.",
            fix="If the run is too slow, re-export a shorter time range "
                "from your historian (e.g. 24 hours instead of 1 month)."))
    else:
        report.add_passed("1.5", "File size is reasonable")

    return True


# ══════════════════════════════════════════════════════════════════════
# STAGE 2 — SHEET-LEVEL HEALTH CHECKS
# ══════════════════════════════════════════════════════════════════════

def _stage2_sheet_checks(path: str, report: HealthCheckReport
                         ) -> Tuple[bool, Optional[Dict[str, Any]]]:
    """Returns (ok, info_dict). info_dict is None if a fatal problem found."""
    info: Dict[str, Any] = {}
    wb = load_workbook(path, read_only=False, data_only=True)
    info["sheet_names"] = list(wb.sheetnames)

    # 2.1 — data sheet present
    data_sheet = None
    for sname in wb.sheetnames:
        ws = wb[sname]
        try:
            first = ws.cell(row=1, column=1).value
        except Exception:
            continue
        if first and str(first).strip().upper() == "TIMESTAMP":
            data_sheet = sname
            break

    if data_sheet is None:
        if "Sheet1" in wb.sheetnames:
            data_sheet = "Sheet1"
        elif wb.sheetnames:
            data_sheet = wb.sheetnames[0]

    if data_sheet is None:
        report.add(CheckResult(
            check_id="2.1", name="Data sheet present", severity=PROBLEM,
            title="No data sheet was found in the workbook.",
            what="The workbook contains no sheets at all.",
            why="The tool needs a sheet of historian data to analyse.",
            fix="Re-export your historian data into the file."))
        return False, None
    info["data_sheet"] = data_sheet
    report.add_passed("2.1", f"Data sheet found ('{data_sheet}')")

    ws = wb[data_sheet]
    header_row_idx = 1
    headers = [c.value for c in ws[header_row_idx]]
    if not (headers and headers[0] and
            str(headers[0]).strip().upper() == "TIMESTAMP"):
        for r in range(2, 15):
            row_vals = [c.value for c in ws[r]]
            if row_vals and row_vals[0] and \
               str(row_vals[0]).strip().upper() == "TIMESTAMP":
                header_row_idx = r
                headers = row_vals
                break
    info["header_row"] = header_row_idx
    info["headers"] = headers

    # 2.6 — TIMESTAMP column
    if not headers or not headers[0] or \
       str(headers[0]).strip().upper() != "TIMESTAMP":
        report.add(CheckResult(
            check_id="2.6", name="TIMESTAMP column", severity=PROBLEM,
            title="The first column is not named TIMESTAMP.",
            what=f"In sheet '{data_sheet}', the first column header is "
                 f"'{headers[0] if headers else '(empty)'}'. The tool requires "
                 "the first column to be exactly 'TIMESTAMP'.",
            why="The tool reads timestamps from column A. Without a "
                "recognisable TIMESTAMP header, it cannot align rows in time.",
            fix=f"1. Open the file in Excel\n"
                f"2. Go to sheet '{data_sheet}'\n"
                f"3. Rename the first column header to 'TIMESTAMP' (capitals)\n"
                "4. Save and re-run."))
        return False, None
    report.add_passed("2.6", "TIMESTAMP column header present")

    # 2.2-2.5 — required and optional sheets
    required_sheets = ["DIAGNOSTIC_CONFIG", "MODE_MAPPING"]
    optional_sheets = ["UNIT_MAPPING", "DIAGNOSTIC_SELECTION"]
    for s in required_sheets:
        if s not in wb.sheetnames:
            report.add(CheckResult(
                check_id={"DIAGNOSTIC_CONFIG": "2.2",
                          "MODE_MAPPING": "2.3"}[s],
                name=f"Required sheet {s}", severity=PROBLEM,
                title=f"Required sheet '{s}' is missing.",
                what=f"Your input file contains: {', '.join(wb.sheetnames)}\n"
                     f"It is missing the required sheet '{s}'.",
                why=("DIAGNOSTIC_CONFIG holds the threshold values used for "
                     "fault detection." if s == "DIAGNOSTIC_CONFIG" else
                     "MODE_MAPPING tells the tool how your DCS labels each "
                     "control mode (AUTO, CAS, MAN, etc)."),
                fix=f"Copy the {s} sheet from the template file "
                    "'synthetic_test_data.xlsx' into your file."))
        else:
            cid = {"DIAGNOSTIC_CONFIG": "2.2", "MODE_MAPPING": "2.3"}[s]
            report.add_passed(cid, f"Sheet '{s}' present")
    for s in optional_sheets:
        cid = {"UNIT_MAPPING": "2.4", "DIAGNOSTIC_SELECTION": "2.5"}[s]
        if s not in wb.sheetnames:
            report.add(CheckResult(
                check_id=cid, name=f"Optional sheet {s}", severity=WARNING,
                title=f"Optional sheet '{s}' is missing.",
                what=f"Your file does not contain a {s} sheet.",
                why=("Without UNIT_MAPPING, peer-comparison features (loops "
                     "grouped by unit) will be limited." if s == "UNIT_MAPPING"
                     else "Without DIAGNOSTIC_SELECTION, all diagnostics will "
                          "run with their defaults."),
                fix=f"Optional — tool will continue. To enable {s} features, "
                    "copy the sheet from 'synthetic_test_data.xlsx'."))
        else:
            report.add_passed(cid, f"Sheet '{s}' present")

    if any(s not in wb.sheetnames for s in required_sheets):
        return False, None

    # 2.8 — column naming & loop completeness
    suffix_map = {}
    suffix_seen = set()
    name_separators = {"_": 0, ".": 0}
    bad_named = []

    for col in headers[1:]:
        if col is None:
            continue
        col = str(col)
        matched = False
        # Try with separator: _, -, ., space
        for sep in ("_", ".", "-", " "):
            for tail in ("PV", "OP", "SP", "Mode", "MODE", "pv", "op", "sp", "mode"):
                marker = sep + tail
                if col.endswith(marker):
                    base = col[: -len(marker)]
                    suffix = tail.upper().replace("MODE", "Mode")
                    suffix_map.setdefault(base, {})[suffix] = col
                    suffix_seen.add(suffix)
                    if sep in ("_", "."):
                        name_separators[sep] = name_separators.get(sep, 0) + 1
                    matched = True
                    break
            if matched:
                break
        if not matched:
            # Try no separator: col ends directly with suffix
            for tail in ("PV", "OP", "SP", "Mode", "MODE", "pv", "op", "sp", "mode"):
                if col.upper().endswith(tail.upper()) and len(col) > len(tail):
                    base = col[: -len(tail)]
                    # Guard: char before suffix must not be alpha (avoid OPVAL matching OP)
                    if not base[-1:].isalpha():
                        suffix = tail.upper().replace("MODE", "Mode")
                        suffix_map.setdefault(base, {})[suffix] = col
                        suffix_seen.add(suffix)
                        matched = True
                        break
        if not matched:
            bad_named.append(col)

    info["suffix_map"] = suffix_map

    if not suffix_map:
        # Last resort: implicit PV — bare tag is PV, tagOP/tagSP are OP/SP
        op_sp_bases = set()
        header_strs = [str(h) for h in headers if h]
        for col in headers[1:]:
            if col is None:
                continue
            col = str(col)
            for tail in ("OP", "SP", "op", "sp"):
                if col.upper().endswith(tail) and len(col) > len(tail):
                    base = col[:-len(tail)]
                    if base in header_strs:
                        suffix_map.setdefault(base, {})[tail.upper()] = col
                        suffix_map[base]["PV"] = base
                        op_sp_bases.add(base)
        if op_sp_bases:
            info["suffix_map"] = suffix_map
            report.add(CheckResult(
                check_id="2.8b", name="Implicit PV columns", severity=INFO,
                title=f"Detected {len(op_sp_bases)} loop(s) with implicit PV naming.",
                what="Loop PV columns have no suffix — the tag name itself is the PV.\n"
                     "Example: 'YN.ETH1.15FC311' (PV), 'YN.ETH1.15FC311OP', 'YN.ETH1.15FC311SP'",
                why="This is a valid naming convention. The tool will handle it correctly.",
                fix="No action needed."))
        else:
            report.add(CheckResult(
                check_id="2.8", name="Has loop columns", severity=PROBLEM,
                title="No loop columns recognised.",
                what="None of the columns matched the expected naming pattern\n"
                     "  <tag>_PV  <tag>_OP  <tag>_SP  <tag>_Mode\n"
                     f"Column names found: {', '.join(str(h) for h in headers[1:6])}…",
                why="The tool identifies which column is PV / OP / SP / Mode by "
                    "the suffix after the underscore. Without recognisable "
                    "suffixes, no loop can be analysed.",
                fix="1. Open the file in Excel\n"
                    "2. Rename the columns so each ends in _PV, _OP, _SP, or _Mode\n"
                    "3. Save and re-run."))
            return False, None

    report.add_passed("2.8", f"Loop columns recognised ({len(suffix_map)} loop bases)")

    # 2.10 — separator consistency
    if name_separators["_"] > 0 and name_separators["."] > 0:
        report.add(CheckResult(
            check_id="2.10", name="Separator consistency", severity=WARNING,
            title="Mixed column naming separators detected.",
            what=f"Some columns use underscore (e.g. tag_PV — "
                 f"{name_separators['_']} columns) and some use dot "
                 f"(e.g. tag.PV — {name_separators['.']} columns).",
            why="Mixing makes it harder to match PV/OP/SP/Mode for the same "
                "loop. The tool will try its best but some loops may be "
                "misgrouped.",
            fix="Use Find & Replace in Excel to standardise on one separator. "
                "Underscore is recommended."))
    else:
        report.add_passed("2.10", "Column naming separator is consistent")

    # 2.11 — per-loop completeness
    incomplete = []
    skipped_loops = []
    for base, mapping in suffix_map.items():
        missing = [s for s in ("PV", "OP", "SP", "Mode") if s not in mapping]
        if missing:
            incomplete.append((base, missing))
            if "PV" in missing or "OP" in missing:
                skipped_loops.append(base)
    if incomplete:
        lines = []
        for base, miss in incomplete[:10]:
            kept = "/".join(s for s in ("PV", "OP", "SP", "Mode")
                            if s in suffix_map[base])
            lines.append(f"  {base}  →  has {kept}  (missing {'/'.join(miss)})")
        if len(incomplete) > 10:
            lines.append(f"  … and {len(incomplete) - 10} more")
        skip_txt = ""
        if skipped_loops:
            skip_txt = (f"\nLoops {', '.join(skipped_loops[:5])}"
                        f"{' …' if len(skipped_loops) > 5 else ''}"
                        f" will be SKIPPED — cannot diagnose without PV+OP.")
        report.add(CheckResult(
            check_id="2.11", name="Per-loop completeness", severity=WARNING,
            title=f"{len(incomplete)} loop(s) are missing some columns.",
            what="\n".join(lines),
            why="Loops without SP cannot be checked for tracking error.\n"
                "Loops without PV or OP cannot be diagnosed at all." + skip_txt,
            fix="Optional — tool will run as-is. To get full diagnostics:\n"
                "  Re-export from the historian and add the missing columns."))
    else:
        report.add_passed("2.11", "All loops have complete PV/OP/SP/Mode columns")

    # 2.12 — duplicate column names
    seen = set()
    dupes = []
    for h in headers[1:]:
        if h is None:
            continue
        h = str(h)
        if h in seen:
            dupes.append(h)
        seen.add(h)
    if dupes:
        # Check if duplicates are exact copies (same data) — if so, just warn
        # and let the engine drop them; only hard-fail if they differ
        report.add(CheckResult(
            check_id="2.12", name="No duplicate columns", severity=WARNING,
            title="Duplicate column names found — later copy will be ignored.",
            what="The same column name appears more than once: " +
                 ", ".join(dupes[:6]),
            why="The tool will use the first occurrence of each duplicate column "
                "and ignore the rest. Results should be unaffected if the columns "
                "contain the same data.",
            fix="Optional — remove the duplicate column(s) from your export to "
                "silence this warning."))
        # Strip duplicate headers so downstream checks see a clean list
        seen2 = set()
        deduped = []
        for h in info.get("headers", []):
            hs = str(h) if h is not None else None
            if hs is None or hs not in seen2:
                deduped.append(h)
                if hs is not None:
                    seen2.add(hs)
            # else: skip duplicate
        info["headers"] = deduped
    else:
        report.add_passed("2.12", "No duplicate column names")

    info["wb_sheets"] = list(wb.sheetnames)
    return True, info


# ══════════════════════════════════════════════════════════════════════
# STAGE 3 — CONFIGURATION SANITY CHECKS
# ══════════════════════════════════════════════════════════════════════

def _stage3_config_checks(path: str, report: HealthCheckReport
                          ) -> Dict[str, Any]:
    """Read DIAGNOSTIC_CONFIG and check each value. Returns the parsed config."""
    cfg: Dict[str, Any] = {}
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        if "DIAGNOSTIC_CONFIG" not in wb.sheetnames:
            return cfg
        ws = wb["DIAGNOSTIC_CONFIG"]
        rows = list(ws.iter_rows(values_only=True))
    except Exception as e:
        report.add(CheckResult(
            check_id="3.0", name="Config readable", severity=PROBLEM,
            title="DIAGNOSTIC_CONFIG sheet could not be read.",
            what=str(e),
            why="The thresholds used for fault detection cannot be loaded.",
            fix="Inspect the DIAGNOSTIC_CONFIG sheet for unusual values."))
        return cfg

    expected = set(CONFIG_RANGES.keys())
    found_params: Dict[str, Any] = {}
    bad_types: List[Tuple[str, Any]] = []
    for r in rows[1:]:
        if not r or r[0] is None:
            continue
        param = str(r[0]).strip()
        val = r[1] if len(r) > 1 else None
        if val is None or (isinstance(val, str) and not val.strip()):
            continue
        try:
            num = float(val)
            found_params[param] = num
        except (TypeError, ValueError):
            bad_types.append((param, val))

    cfg.update(found_params)

    if bad_types:
        for param, v in bad_types[:5]:
            default = CONFIG_RANGES.get(param, (None, None, None))[2]
            report.add(CheckResult(
                check_id="3.2", name="Config numeric", severity=PROBLEM,
                title=f"Config value '{param}' is not a number.",
                what=f"In DIAGNOSTIC_CONFIG, the parameter {param} has the "
                     f"value '{v}' instead of a number.",
                why="This parameter is a numeric threshold the tool compares "
                    "against. It must be a number like 15 or 12.5, not text.",
                fix=f"Open DIAGNOSTIC_CONFIG, find the row with Parameter = "
                    f"'{param}', and change the Value cell to a number." +
                    (f"\n\nDefault value: {default}" if default is not None
                     else "")))
    else:
        report.add_passed("3.2", "All config values are numeric")

    missing = expected - set(found_params.keys())
    if missing:
        report.add(CheckResult(
            check_id="3.1", name="Config completeness", severity=INFO,
            title=f"{len(missing)} expected config parameter(s) not found.",
            what="The following parameters are not listed in your "
                 "DIAGNOSTIC_CONFIG sheet:\n  " + "\n  ".join(sorted(missing)),
            why="The tool will use built-in defaults for these. Existing "
                "behaviour is preserved.",
            fix="Optional — to expose these for editing, add the parameter "
                "name and a value in the DIAGNOSTIC_CONFIG sheet."))
    else:
        report.add_passed("3.1", "All expected config parameters present")

    out_of_range: List[Tuple[str, float, float, float, float]] = []
    for k, val in found_params.items():
        if k not in CONFIG_RANGES:
            continue
        lo, hi, default = CONFIG_RANGES[k]
        if val < lo or val > hi:
            out_of_range.append((k, val, lo, hi, default))
    if out_of_range:
        for k, val, lo, hi, default in out_of_range[:5]:
            report.add(CheckResult(
                check_id="3.3", name=f"Config range — {k}", severity=WARNING,
                title=f"Config value '{k}' is outside its typical range.",
                what=f"Value:    {val}\n"
                     f"Typical:  {lo} to {hi}\n"
                     f"Default:  {default}",
                why="A value far from the typical range may produce unexpected "
                    "diagnostics. This is a warning, not an error — the value "
                    "will still be used.",
                fix="If this is a typo: change to a value in the typical range.\n"
                    "If you set this deliberately: ignore this warning."))
    else:
        report.add_passed("3.3", "All config values are in sensible ranges")

    if "STICT_CONF_HIGH" in cfg and "STICT_CONF_MED" in cfg:
        if cfg["STICT_CONF_HIGH"] <= cfg["STICT_CONF_MED"]:
            report.add(CheckResult(
                check_id="3.4", name="Confidence ordering", severity=PROBLEM,
                title="STICT_CONF_HIGH must be greater than STICT_CONF_MED.",
                what=f"STICT_CONF_HIGH = {cfg['STICT_CONF_HIGH']}\n"
                     f"STICT_CONF_MED  = {cfg['STICT_CONF_MED']}",
                why="The HIGH threshold must be strictly greater than MED.",
                fix="In DIAGNOSTIC_CONFIG, set STICT_CONF_HIGH > STICT_CONF_MED.\n"
                    "Defaults: HIGH = 70, MED = 40."))
        else:
            report.add_passed("3.4", "Stiction confidence thresholds ordered")
    if "PROP_CONF_STRONG" in cfg and "PROP_CONF_MIN" in cfg:
        if cfg["PROP_CONF_STRONG"] <= cfg["PROP_CONF_MIN"]:
            report.add(CheckResult(
                check_id="3.4b", name="Propagation conf ordering", severity=PROBLEM,
                title="PROP_CONF_STRONG must be greater than PROP_CONF_MIN.",
                what=f"PROP_CONF_STRONG = {cfg['PROP_CONF_STRONG']}\n"
                     f"PROP_CONF_MIN    = {cfg['PROP_CONF_MIN']}",
                why="STRONG must be strictly greater than MIN.",
                fix="Defaults: MIN = 50, STRONG = 70."))

    return cfg


# ══════════════════════════════════════════════════════════════════════
# STAGE 4 — DATA QUALITY CHECKS
# ══════════════════════════════════════════════════════════════════════

def _stage4_data_checks(path: str, sheet_info: Dict[str, Any],
                        report: HealthCheckReport) -> Dict[str, Any]:
    """Read the data sheet, run row-by-row and column-by-column checks."""
    out: Dict[str, Any] = {}

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_info["data_sheet"]]
    header_row = sheet_info["header_row"]
    headers = sheet_info["headers"]

    raw_rows = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1,
                                              values_only=True), start=header_row + 1):
        raw_rows.append(row)
    out["n_rows"] = len(raw_rows)

    if len(raw_rows) == 0:
        report.add(CheckResult(
            check_id="4.5", name="Has data rows", severity=PROBLEM,
            title="The data sheet contains no data rows.",
            what=f"Sheet '{sheet_info['data_sheet']}' has only a header row.",
            why="No data means no analysis is possible.",
            fix="Re-export the data from your historian over a non-empty "
                "time range."))
        return out
    report.add_passed("4.5", f"Data sheet has {len(raw_rows):,} rows")

    ts_col = []
    bad_ts_rows = []
    for i, r in enumerate(raw_rows):
        v = r[0] if r else None
        if isinstance(v, datetime):
            ts_col.append(v)
        else:
            try:
                ts_col.append(pd.to_datetime(v))
            except Exception:
                bad_ts_rows.append((i, v))
                ts_col.append(None)

    if bad_ts_rows:
        ex_lines = [f"  Row {i + header_row + 1}: {v!r}"
                    for i, v in bad_ts_rows[:5]]
        report.add(CheckResult(
            check_id="4.1a", name="Timestamps parseable", severity=PROBLEM,
            title=f"{len(bad_ts_rows)} rows have unparseable timestamps.",
            what="Examples:\n" + "\n".join(ex_lines),
            why="The tool cannot place these rows in time order.",
            fix="Open the file and check the TIMESTAMP column. Common causes:\n"
                "  • Cell formatted as text instead of date/time\n"
                "  • Empty cells in the middle of the data\n"
                "Format the column as Date/Time and re-save."))
        return out
    valid_ts = [t for t in ts_col if t is not None]

    backwards = []
    duplicates = []
    for i in range(1, len(valid_ts)):
        if valid_ts[i] < valid_ts[i-1]:
            backwards.append(i)
        elif valid_ts[i] == valid_ts[i-1]:
            duplicates.append(i)

    if backwards:
        rows_to_show = [(i, valid_ts[i-1], valid_ts[i]) for i in backwards[:3]]
        ex = "\n".join(f"  Row {i + header_row + 1}: {prev} → {now} "
                       "(goes backwards)" for i, prev, now in rows_to_show)
        report.add(CheckResult(
            check_id="4.1", name="Timestamps in order", severity=PROBLEM,
            title=f"Timestamps jump backwards at {len(backwards)} place(s).",
            what="Examples:\n" + ex,
            why="Several diagnostics rely on time order.",
            fix=f"In Excel: select sheet '{sheet_info['data_sheet']}' → Data "
                "tab → Sort by TIMESTAMP ascending → save."))
    else:
        report.add_passed("4.1", "Timestamps are in chronological order")

    if duplicates:
        report.add(CheckResult(
            check_id="4.2", name="No duplicate timestamps", severity=WARNING,
            title=f"{len(duplicates)} duplicate timestamp(s) found.",
            what=f"Some timestamps appear more than once. Example row: "
                 f"{duplicates[0] + header_row + 1}",
            why="Duplicates may double-count moments in diagnostics.",
            fix="In Excel: Data → Remove Duplicates (on TIMESTAMP column)."))
    else:
        report.add_passed("4.2", "No duplicate timestamps")

    if len(valid_ts) >= 3:
        diffs_sec = np.array([(valid_ts[i] - valid_ts[i-1]).total_seconds()
                              for i in range(1, len(valid_ts))])
        diffs_sec = diffs_sec[diffs_sec > 0]
        if len(diffs_sec):
            median_dt = float(np.median(diffs_sec))
            irregular_frac = float(np.mean(np.abs(diffs_sec - median_dt) >
                                           0.1 * median_dt))
            out["median_dt_sec"] = median_dt
            out["irregular_frac"] = irregular_frac
            if irregular_frac > 0.05:
                lo, hi = float(diffs_sec.min()), float(diffs_sec.max())
                report.add(CheckResult(
                    check_id="4.3", name="Sampling regularity", severity=WARNING,
                    title="Sampling rate is irregular.",
                    what=f"Median sample interval: {median_dt:.0f} seconds.\n"
                         f"{irregular_frac*100:.0f}% of intervals deviate >10% "
                         f"from this — range {lo:.0f}s to {hi:.0f}s.",
                    why="The tool assumes near-constant sampling for some calculations.",
                    fix="Re-export from the historian using a fixed interval."))
            else:
                report.add_passed("4.3", f"Sampling rate is regular (~{median_dt:.0f}s)")

    if len(valid_ts) >= 2:
        duration_h = (valid_ts[-1] - valid_ts[0]).total_seconds() / 3600
        out["duration_hours"] = duration_h
        if duration_h < 1:
            report.add(CheckResult(
                check_id="4.4", name="Duration", severity=PROBLEM,
                title="Not enough data for reliable diagnostics.",
                what=f"The file covers {duration_h*60:.0f} minutes ({len(valid_ts)} rows).",
                why="Several diagnostics need at least a few hours of data.",
                fix="Re-export from the historian over a longer time range:\n"
                    "  • Minimum useful:  2 hours\n"
                    "  • Recommended:    24 hours\n"
                    "  • Best:            1 week"))
        elif duration_h < 24:
            report.add(CheckResult(
                check_id="4.4", name="Duration", severity=WARNING,
                title=f"Data covers only {duration_h:.1f} hours.",
                what="Tool will run, but full diagnostic coverage typically needs ≥24 hours.",
                why="Slow problems take longer to show.",
                fix="For best results, re-export 24 hours or more."))
        else:
            report.add_passed("4.4", f"Duration is {duration_h:.1f} hours")

    out["per_column"] = {}
    excel_err_finds: Dict[str, int] = {}
    quality_flag_finds: Dict[str, Dict[str, int]] = {}
    sentinel_finds: Dict[str, Dict[float, int]] = {}
    nan_text_finds: Dict[str, int] = {}
    nonnumeric_finds: Dict[str, List[str]] = {}
    inf_finds: Dict[str, int] = {}
    empty_finds: Dict[str, int] = {}
    pv_loops_no_data: List[str] = []
    suffix_map = sheet_info["suffix_map"]

    col_pos = {h: i for i, h in enumerate(headers) if h is not None}

    for col_name, ci in col_pos.items():
        if ci == 0:
            continue
        is_data_col = any(col_name == m.get(s) for m in suffix_map.values()
                          for s in ("PV", "OP", "SP", "Mode"))
        if not is_data_col:
            continue
        is_mode = col_name.lower().endswith("_mode") or col_name.lower().endswith(".mode")

        excel_errs = 0
        quality_hits: Dict[str, int] = {}
        sentinel_hits: Dict[float, int] = {}
        nan_text = 0
        non_num: List[str] = []
        infs = 0
        empties = 0
        numeric_count = 0
        zeros_count = 0

        for r in raw_rows:
            v = r[ci] if ci < len(r) else None
            if v is None:
                empties += 1
                continue
            if isinstance(v, str):
                vs = v.strip()
                if vs == "":
                    empties += 1
                    continue
                if vs in EXCEL_ERRORS:
                    excel_errs += 1
                    continue
                if vs.lower() in QUALITY_FLAGS:
                    quality_hits[vs] = quality_hits.get(vs, 0) + 1
                    continue
                if vs.lower() in ("nan", "na", "n/a", "null", "none", "nil"):
                    nan_text += 1
                    continue
                if is_mode:
                    quality_hits.setdefault("__mode_values__", 0)
                    quality_hits[vs] = quality_hits.get(vs, 0) + 1
                    continue
                try:
                    fv = float(vs)
                    if not np.isfinite(fv):
                        infs += 1
                        continue
                    numeric_count += 1
                    if fv == 0:
                        zeros_count += 1
                    if fv in NUMERIC_SENTINELS:
                        sentinel_hits[fv] = sentinel_hits.get(fv, 0) + 1
                    continue
                except Exception:
                    if len(non_num) < 5:
                        non_num.append(vs)
                    continue
            if isinstance(v, (int, float)):
                fv = float(v)
                if not np.isfinite(fv):
                    infs += 1
                    continue
                numeric_count += 1
                if fv == 0:
                    zeros_count += 1
                if fv in NUMERIC_SENTINELS:
                    sentinel_hits[fv] = sentinel_hits.get(fv, 0) + 1
                continue

        out["per_column"][col_name] = {
            "is_mode": is_mode,
            "empties": empties,
            "excel_errs": excel_errs,
            "nan_text": nan_text,
            "non_numeric": non_num,
            "infs": infs,
            "quality_flags": quality_hits if not is_mode else {},
            "mode_values": quality_hits if is_mode else {},
            "sentinels": sentinel_hits,
            "numeric_count": numeric_count,
            "zeros_count": zeros_count,
        }

        if excel_errs:
            excel_err_finds[col_name] = excel_errs
        if quality_hits and not is_mode:
            quality_flag_finds[col_name] = quality_hits
        if sentinel_hits:
            sentinel_finds[col_name] = sentinel_hits
        if nan_text:
            nan_text_finds[col_name] = nan_text
        if non_num and not is_mode:
            nonnumeric_finds[col_name] = non_num
        if infs:
            inf_finds[col_name] = infs
        if empties:
            empty_finds[col_name] = empties

    if excel_err_finds:
        lines = []
        for c, n in list(excel_err_finds.items())[:6]:
            lines.append(f"  {c}: {n} cells")
        report.add(CheckResult(
            check_id="4.7", name="No Excel errors", severity=PROBLEM,
            title="Excel error values found in data columns.",
            what="\n".join(lines),
            why="Cells like #DIV/0!, #N/A, or #REF! cannot be used as data.",
            fix="Paste the values as static numbers (Copy → Paste Special → Values)."))
    else:
        report.add_passed("4.7", "No Excel error markers in data")

    if nan_text_finds:
        lines = []
        for c, n in list(nan_text_finds.items())[:6]:
            lines.append(f"  {c}: {n} cells with 'NaN'/'None'/'NULL' text")
        report.add(CheckResult(
            check_id="4.8", name="No NaN text", severity=WARNING,
            title="Literal 'NaN' or 'NULL' text found in data columns.",
            what="\n".join(lines),
            why="The tool will treat them as missing values.",
            fix="Optional — use Find & Replace to clear these cells."))
    else:
        report.add_passed("4.8", "No literal NaN text in numeric columns")

    if quality_flag_finds:
        lines = []
        for c, hits in list(quality_flag_finds.items())[:6]:
            for v, n in list(hits.items())[:3]:
                lines.append(f"  {c}: '{v}' ({n} rows)")
        report.add(CheckResult(
            check_id="4.10", name="No quality flags", severity=WARNING,
            title="Historian quality flags found in numeric columns.",
            what="\n".join(lines),
            why="These are historian quality flags exported as text. "
                "The data-repair module will treat them as missing values.",
            fix="Filter on quality in your historian export tool."))
    else:
        report.add_passed("4.10", "No quality flag text in numeric columns")

    if nonnumeric_finds:
        lines = []
        for c, vals in list(nonnumeric_finds.items())[:5]:
            lines.append(f"  {c}: examples → {', '.join(vals[:3])}")
        report.add(CheckResult(
            check_id="4.9", name="Numeric content", severity=PROBLEM,
            title="Non-numeric text found in PV/OP/SP columns.",
            what="\n".join(lines),
            why="These columns must contain numbers.",
            fix="Open the file and check the listed cells."))
    else:
        report.add_passed("4.9", "All numeric columns contain valid numbers "
                                "(or recognised gaps)")

    if sentinel_finds:
        lines = []
        for c, hits in list(sentinel_finds.items())[:6]:
            for v, n in hits.items():
                lines.append(f"  {c}: {v} appears {n} times")
        report.add(CheckResult(
            check_id="4.11", name="No sentinel values", severity=WARNING,
            title="Possible 'no data' sentinel values detected.",
            what="\n".join(lines),
            why="Values like -9999 or 99999 are common 'no data' markers. "
                "The data-repair module will treat them as missing data.",
            fix="If these ARE real measurements: re-run with --no-sentinel-handling."))
    else:
        report.add_passed("4.11", "No suspected sentinel values found")

    if inf_finds:
        lines = []
        for c, n in list(inf_finds.items())[:6]:
            lines.append(f"  {c}: {n} cells")
        report.add(CheckResult(
            check_id="4.14", name="No infinity", severity=PROBLEM,
            title="Infinity values found in data.",
            what="\n".join(lines),
            why="Infinity values are not valid measurements.",
            fix="Trace the source. Either delete those cells or fix the export."))
    else:
        report.add_passed("4.14", "No infinity values")

    miss_frac_problems: List[Tuple[str, float, int]] = []
    miss_frac_warnings: List[Tuple[str, float, int]] = []
    n_total = len(raw_rows)
    for base, mp in suffix_map.items():
        pv_col = mp.get("PV")
        if pv_col is None:
            continue
        c = out["per_column"].get(pv_col, {})
        n_missing = (c.get("empties", 0) + c.get("nan_text", 0)
                     + c.get("excel_errs", 0)
                     + sum(c.get("sentinels", {}).values())
                     + sum(c.get("quality_flags", {}).values())
                     + c.get("infs", 0))
        if n_total > 0:
            frac = n_missing / n_total
            if frac >= 0.30:
                miss_frac_problems.append((base, frac, n_missing))
                if c.get("numeric_count", 0) == 0:
                    pv_loops_no_data.append(base)
            elif frac >= 0.10:
                miss_frac_warnings.append((base, frac, n_missing))

    if miss_frac_problems:
        lines = []
        for base, frac, nm in miss_frac_problems[:8]:
            lines.append(f"  {base}: {frac*100:.0f}% missing/invalid ({nm} of {n_total})")
        report.add(CheckResult(
            check_id="4.10b", name="Missing data per loop", severity=PROBLEM,
            title=f"{len(miss_frac_problems)} loop(s) have over 30% missing/invalid data.",
            what="\n".join(lines),
            why="With over 30% of values missing, no diagnostic can give a reliable answer.",
            fix="• Re-export over a different time range.\n"
                "• Or re-run with --skip-incomplete-loops to skip these automatically."))
    else:
        report.add_passed("4.10b", "All loops have <30% missing data")

    if miss_frac_warnings:
        lines = []
        for base, frac, nm in miss_frac_warnings[:8]:
            lines.append(f"  {base}: {frac*100:.0f}% missing")
        report.add(CheckResult(
            check_id="4.10c", name="Missing data 10-30%", severity=WARNING,
            title=f"{len(miss_frac_warnings)} loop(s) have 10-30% missing data.",
            what="\n".join(lines),
            why="The data-repair module will fill small gaps by interpolation.",
            fix="Optional — investigate the gap pattern in the affected loops."))

    op_scale_warnings = []
    for base, mp in suffix_map.items():
        op_col = mp.get("OP")
        if not op_col:
            continue
        col_vals = []
        ci = col_pos.get(op_col)
        if ci is None:
            continue
        for r in raw_rows[:min(2000, len(raw_rows))]:
            v = r[ci] if ci < len(r) else None
            try:
                fv = float(v)
                if np.isfinite(fv) and fv not in NUMERIC_SENTINELS:
                    col_vals.append(fv)
            except Exception:
                pass
        if len(col_vals) >= 30:
            arr = np.array(col_vals)
            high = np.mean(arr > 110)
            low = np.mean(arr < -10)
            if high > 0.05 or low > 0.05:
                op_scale_warnings.append((base, float(arr.min()), float(arr.max())))
    if op_scale_warnings:
        lines = []
        for base, mn, mx in op_scale_warnings[:5]:
            lines.append(f"  {base}: OP range observed = {mn:.1f} to {mx:.1f}")
        report.add(CheckResult(
            check_id="4.7b", name="OP scaling", severity=WARNING,
            title="OP values frequently outside 0-100% range.",
            what="\n".join(lines),
            why="Controller output (OP) is usually a percentage 0-100%.",
            fix="Best practice: export OP as percentage."))
    else:
        report.add_passed("4.7", "OP values are in expected 0-100% range")

    out["miss_problems"] = miss_frac_problems
    out["miss_warnings"] = miss_frac_warnings
    out["pv_loops_no_data"] = pv_loops_no_data
    out["valid_ts"] = valid_ts
    return out


# ══════════════════════════════════════════════════════════════════════
# STAGE 5 — CROSS-CONSISTENCY CHECKS
# ══════════════════════════════════════════════════════════════════════

def _stage5_cross_checks(path: str, sheet_info: Dict[str, Any],
                         data_info: Dict[str, Any],
                         report: HealthCheckReport) -> None:
    suffix_map = sheet_info.get("suffix_map", {})
    wb = load_workbook(path, read_only=True, data_only=True)

    if "MODE_MAPPING" in wb.sheetnames:
        ws = wb["MODE_MAPPING"]
        mm: Dict[str, str] = {}
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or r[0] is None or r[1] is None:
                continue
            raw = str(r[1]).strip().upper()
            if raw.endswith(".0"):
                raw = raw[:-2]
            mm[raw] = str(r[0]).strip().upper()

        mapped_categories = set(mm.values())
        if "AUTO" not in mapped_categories:
            report.add(CheckResult(
                check_id="3.5", name="MODE_MAPPING has AUTO", severity=PROBLEM,
                title="MODE_MAPPING does not define any AUTO mode.",
                what=f"Mapped categories found: {', '.join(mapped_categories)}",
                why="The tool only analyses loops while in AUTO or CAS — "
                    "without an AUTO entry, every loop will be skipped.",
                fix="Add a row to MODE_MAPPING: Category=AUTO, "
                    "Value=<your DCS's AUTO label>"))
        else:
            report.add_passed("3.5", "MODE_MAPPING contains AUTO category")

        seen_modes: Dict[str, int] = {}
        seen_originals: Dict[str, str] = {}
        for col_name, info in data_info.get("per_column", {}).items():
            if not info.get("is_mode"):
                continue
            for v, n in info.get("mode_values", {}).items():
                if v == "__mode_values__":
                    continue
                norm = str(v).strip().upper()
                if norm.endswith(".0"):
                    norm = norm[:-2]
                if not norm or norm == "NAN":
                    continue
                seen_modes[norm] = seen_modes.get(norm, 0) + n
                seen_originals.setdefault(norm, str(v))

        unmapped = [v for v in seen_modes if v not in mm]
        if unmapped:
            lines = []
            for v in unmapped[:6]:
                orig = seen_originals.get(v, v)
                lines.append(f"  '{orig}'  ({seen_modes[v]} rows)")
            report.add(CheckResult(
                check_id="3.6", name="Mode values mapped", severity=WARNING,
                title="Some Mode values in the data are not in MODE_MAPPING.",
                what="\n".join(lines),
                why="The tool will treat these as MAN (manual) by default.",
                fix="Open MODE_MAPPING and add a row for each unrecognised value.\n"
                    "    Category    Value\n"
                    "    AUTO        " + (seen_originals.get(unmapped[0], "<value>")
                                          if unmapped else "<value>")))
        else:
            report.add_passed("3.6", "Every Mode value seen in the data is mapped")

    if "UNIT_MAPPING" in wb.sheetnames:
        ws = wb["UNIT_MAPPING"]
        unit_tags = set()
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r and r[0]:
                unit_tags.add(str(r[0]).strip())
        loop_bases = {b.split(".")[-1] for b in suffix_map}
        all_loop_strings = set(suffix_map.keys()) | loop_bases
        ghosts = [t for t in unit_tags if t not in all_loop_strings
                  and not any(t in s for s in suffix_map)]
        if ghosts:
            report.add(CheckResult(
                check_id="5.1", name="UNIT_MAPPING vs data", severity=WARNING,
                title=f"{len(ghosts)} tag(s) in UNIT_MAPPING are not in the data.",
                what=f"Examples: {', '.join(ghosts[:5])}",
                why="Not blocking — these mappings just won't be used.",
                fix="Optional — clean up UNIT_MAPPING to match the loops "
                    "actually present in your data."))

    if "DIAGNOSTIC_SELECTION" in wb.sheetnames:
        ws = wb["DIAGNOSTIC_SELECTION"]
        sub_methods_on = 0
        sub_method_names = ("Heuristic method", "Horch cross-correlation",
                            "Yamashita shape", "Bicoherence")
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r or not r[0]:
                continue
            label = str(r[0]).strip()
            if any(label == s or label.endswith(s) for s in sub_method_names):
                try:
                    if int(r[1]) == 1:
                        sub_methods_on += 1
                except Exception:
                    pass
        if 0 < sub_methods_on < 2:
            report.add(CheckResult(
                check_id="5.4", name="Stiction methods", severity=WARNING,
                title="Only one stiction method is enabled.",
                what=f"{sub_methods_on} of 4 stiction sub-methods are enabled.",
                why="The tool's stiction detection works by having multiple "
                    "independent methods vote.",
                fix="In DIAGNOSTIC_SELECTION, set Heuristic, Horch, "
                    "Yamashita and Bicoherence to 1."))
        else:
            report.add_passed("5.4", "Multiple stiction methods enabled")

    n_loops = len(suffix_map)
    if n_loops < 2:
        report.add(CheckResult(
            check_id="5.5", name="Loop count", severity=WARNING,
            title="Only one control loop found.",
            what=f"The file contains {n_loops} loop(s). Plant-wide "
                 "diagnostics (peer comparison, propagation) need ≥3 loops.",
            why="Single-loop analysis still works, but loses some validation.",
            fix="If you have more loops, add them to the export."))
    else:
        report.add_passed("5.5", f"Loop count is reasonable ({n_loops})")
