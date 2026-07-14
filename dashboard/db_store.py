"""
Postgres-backed data store for the Valve Diagnostic Tool POC.

Every Excel upload gets melted into long format (tag_name, ts, value/text_value)
and upserted into Postgres. The "Run from Database" tab queries a date range
across selected loops, pivots it back into the exact wide Sheet1 layout the
engine already expects, and writes it to a temp .xlsx — so run_v3() and the
reader/engine pipeline are never touched, we just give them a different
source file.

Column type (numeric vs text, e.g. MODE columns) is decided per-column by
sampling values, not by a hardcoded suffix list — works with any tag
naming scheme (PV/OP/SP/MODE, MV/CV/SV/AOUT, etc.) the same way the rest
of the app already handles dynamic suffixes.
"""

import json
import os
from datetime import datetime
from pathlib import Path
from typing import Optional

import asyncpg
import openpyxl

_pool: Optional[asyncpg.Pool] = None

_DEFAULT_CONFIG_PATH = Path(__file__).parent / "default_config.json"

_SEPARATORS = ["_", "-", ".", " "]


def _load_required_sheet_rows():
    """Read DIAGNOSTIC_CONFIG and MODE_MAPPING data from default_config.json
    so the engine's health check (which hard-requires both sheets to be
    present) passes on DB-rebuilt files. The actual threshold values used
    by the engine still come from the config dict param passed to run_v3()
    — these sheets exist only to satisfy the health check's presence/shape
    checks, so shipping the defaults here is correct regardless of what
    the user's saved config overrides."""
    fallback_diag_config = [
        ("AMP_THRESHOLD", 15), ("OP_ACTIVITY_THRESHOLD", 1.5),
        ("IAE_PER_HOUR_THRESHOLD", 200), ("STICT_CONF_HIGH", 70),
        ("STICT_CONF_MED", 40), ("PROP_CONF_MIN", 60),
        ("PROP_CONF_STRONG", 70), ("SERVICE_FACTOR_MIN_PCT", 70),
        ("SS_DETECTION_WINDOW", 30), ("SS_STD_THRESHOLD", 0.5),
        ("FROZEN_SAMPLES_MIN", 10), ("QUANTISATION_UNIQUE_VALS_MAX", 20),
        ("COMPRESSION_FLAT_FRACTION_MAX", 0.3), ("OSCILLATION_REGULARITY_MIN", 0.6),
        ("STICTION_S_MIN_PCT", 0.5), ("HARRIS_INDEX_THRESHOLD", 0.3),
    ]
    fallback_mode_mapping = [
        ("AUTO", "AUTO"), ("AUTO", "A"), ("AUTO", "1"), ("AUTO", "Automatic"),
        ("CAS", "CAS"), ("CAS", "C"), ("CAS", "2"), ("CAS", "Cascade"),
        ("RCAS", "RCAS"), ("RCAS", "R"), ("RCAS", "3"), ("RCAS", "Remote"),
        ("MAN", "MAN"), ("MAN", "M"), ("MAN", "0"), ("MAN", "Manual"),
    ]
    try:
        with open(_DEFAULT_CONFIG_PATH, "r", encoding="utf-8") as f:
            cfg = json.load(f)
        diag_config = [(r["parameter"], r["value"]) for r in cfg.get("diagnostic_config", [])] \
            or fallback_diag_config
        mode_mapping = [(r["category"], r["value"]) for r in cfg.get("mode_mapping", [])] \
            or fallback_mode_mapping
        return diag_config, mode_mapping
    except Exception:
        return fallback_diag_config, fallback_mode_mapping


async def get_pool() -> asyncpg.Pool:
    global _pool
    if _pool is None:
        database_url = os.getenv("DATABASE_URL", "").strip().strip('"').strip("'")
        if not database_url:
            raise RuntimeError("DATABASE_URL not set in .env")
        _pool = await asyncpg.create_pool(database_url, min_size=1, max_size=5)
    return _pool


async def close_pool():
    global _pool
    if _pool is not None:
        await _pool.close()
        _pool = None


def derive_loop_name(tag_name: str) -> str:
    """Strip the trailing role suffix (PV/OP/SP/MODE/MV/... whatever the
    plant uses) off a tag name to get the loop base name. Generic: splits
    at whichever separator occurs last in the string."""
    last_idx = -1
    for sep in _SEPARATORS:
        idx = tag_name.rfind(sep)
        if idx > last_idx:
            last_idx = idx
    return tag_name[:last_idx] if last_idx > 0 else tag_name


def _is_numeric(v) -> bool:
    if v is None:
        return True
    try:
        float(v)
        return True
    except (TypeError, ValueError):
        return False


async def insert_readings_from_excel(xlsx_path: str, source_file: str, uploaded_by: str) -> int:
    """Read Sheet1, melt to long format, upsert into Postgres. Returns rows written."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return 0

    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    data_rows = rows[1:]

    # Decide numeric vs text per column by sampling up to 200 non-null values
    numeric_col = [True] * len(headers)
    for col_idx in range(1, len(headers)):
        sample = [r[col_idx] for r in data_rows[:200] if col_idx < len(r) and r[col_idx] is not None]
        if sample:
            numeric_col[col_idx] = all(_is_numeric(v) for v in sample)

    records = []
    for row in data_rows:
        ts = row[0]
        if ts is None:
            continue
        if hasattr(ts, "strftime"):
            ts_dt = ts
        else:
            try:
                ts_dt = datetime.fromisoformat(str(ts))
            except ValueError:
                continue

        for col_idx in range(1, len(headers)):
            tag_name = headers[col_idx]
            if not tag_name or col_idx >= len(row):
                continue
            raw_val = row[col_idx]
            if raw_val is None:
                continue
            if numeric_col[col_idx]:
                try:
                    val, text_val = float(raw_val), None
                except (TypeError, ValueError):
                    val, text_val = None, None
            else:
                val, text_val = None, str(raw_val).strip()
            records.append((tag_name, ts_dt, val, text_val, source_file, uploaded_by))

    if not records:
        return 0

    pool = await get_pool()
    async with pool.acquire() as conn:
        await conn.executemany(
            """
            INSERT INTO tag_readings (tag_name, ts, value, text_value, source_file, uploaded_by)
            VALUES ($1, $2, $3, $4, $5, $6)
            ON CONFLICT (tag_name, ts) DO UPDATE
                SET value = EXCLUDED.value,
                    text_value = EXCLUDED.text_value,
                    source_file = EXCLUDED.source_file,
                    uploaded_by = EXCLUDED.uploaded_by,
                    uploaded_at = now()
            """,
            records,
        )
        await _refresh_loop_summary(conn, {r[0] for r in records})
    return len(records)


async def _refresh_loop_summary(conn, touched_tag_names: set):
    """Recompute loop_summary rows for whichever loops this upload touched.
    Scoped to just those loops' tags — a handful of small indexed queries,
    not a full-table scan. Keeps the Database tab's loop list fast to load
    no matter how large tag_readings grows."""
    if not touched_tag_names:
        return
    affected_loops = sorted({derive_loop_name(t) for t in touched_tag_names})
    for loop in affected_loops:
        loop_tags = [t for t in touched_tag_names if derive_loop_name(t) == loop]
        agg = await conn.fetchrow(
            "SELECT COUNT(*) AS cnt, MIN(ts) AS min_ts, MAX(ts) AS max_ts "
            "FROM tag_readings WHERE tag_name = ANY($1::text[])",
            loop_tags,
        )
        await conn.execute(
            """
            INSERT INTO loop_summary (loop_name, reading_count, min_ts, max_ts, updated_at)
            VALUES ($1, $2, $3, $4, now())
            ON CONFLICT (loop_name) DO UPDATE
                SET reading_count = EXCLUDED.reading_count,
                    min_ts = EXCLUDED.min_ts,
                    max_ts = EXCLUDED.max_ts,
                    updated_at = now()
            """,
            loop, agg["cnt"], agg["min_ts"], agg["max_ts"],
        )


async def list_available_loops() -> dict:
    """Loop names + overall date range currently stored. Reads from the
    loop_summary fast-path table; falls back to a full scan only if that
    table is somehow empty (e.g. data existed before this table did)."""
    pool = await get_pool()
    async with pool.acquire() as conn:
        summary_rows = await conn.fetch(
            "SELECT loop_name, min_ts, max_ts FROM loop_summary ORDER BY loop_name"
        )
        if summary_rows:
            mins = [r["min_ts"] for r in summary_rows if r["min_ts"]]
            maxs = [r["max_ts"] for r in summary_rows if r["max_ts"]]
            return {
                "loops": [{"loop": r["loop_name"], "tags": []} for r in summary_rows],
                "min_ts": min(mins).isoformat() if mins else None,
                "max_ts": max(maxs).isoformat() if maxs else None,
            }

        # Fallback: no summary rows yet — do the full scan once so the tab
        # still works, and backfill loop_summary so this path isn't needed again.
        tags = await conn.fetch("SELECT DISTINCT tag_name FROM tag_readings ORDER BY tag_name")
        span = await conn.fetchrow("SELECT MIN(ts) AS min_ts, MAX(ts) AS max_ts FROM tag_readings")
        tag_names = [r["tag_name"] for r in tags]
        if tag_names:
            await _refresh_loop_summary(conn, set(tag_names))

    loops: dict = {}
    for t in tag_names:
        loops.setdefault(derive_loop_name(t), []).append(t)

    return {
        "loops": [{"loop": loop, "tags": tags_} for loop, tags_ in sorted(loops.items())],
        "min_ts": span["min_ts"].isoformat() if span["min_ts"] else None,
        "max_ts": span["max_ts"].isoformat() if span["max_ts"] else None,
    }


async def build_excel_from_db(loop_names: list, start_ts: str, end_ts: str, out_path: str,
                               suffix_map: dict = None) -> int:
    """Query all tags belonging to the given loops between start/end, pivot
    into the wide Sheet1 layout (TIMESTAMP | tag1 | tag2 | ...), write to
    out_path. If suffix_map is given (e.g. {'pv':['PV','MV'], 'sp':['SP'],
    'op':['OP'], 'mode':['MODE']}), also writes a 'loop_format' sheet so
    reader.read_loop_format() picks up the same custom suffixes the user
    defined — needed when a plant doesn't use plain PV/SP/OP/MODE tag names.
    Returns the number of timestamp rows written."""
    if not loop_names:
        raise ValueError("No loops selected")

    # Frontend sends datetime-local strings like '2026-04-27T00:00' —
    # asyncpg needs real datetime objects for a timestamp param, the
    # ::timestamp SQL cast alone doesn't satisfy its client-side type check.
    try:
        start_dt = datetime.fromisoformat(start_ts)
        end_dt = datetime.fromisoformat(end_ts)
    except ValueError:
        raise ValueError(f"Could not parse date range: {start_ts!r} / {end_ts!r}")

    pool = await get_pool()
    async with pool.acquire() as conn:
        all_tags_rows = await conn.fetch("SELECT DISTINCT tag_name FROM tag_readings")
        all_tags = [r["tag_name"] for r in all_tags_rows]
        wanted_tags = [t for t in all_tags if derive_loop_name(t) in loop_names]

        if not wanted_tags:
            raise ValueError("No tags found for the selected loops")

        rows = await conn.fetch(
            """
            SELECT tag_name, ts, value, text_value
            FROM tag_readings
            WHERE tag_name = ANY($1::text[])
              AND ts BETWEEN $2 AND $3
            ORDER BY ts
            """,
            wanted_tags, start_dt, end_dt,
        )

    # Per-loop coverage check — catches both "nothing in range at all" and
    # "some loops have no data in this range" with one clear, named error,
    # instead of silently building a file with blank columns for whichever
    # loop(s) came up empty.
    loops_with_data = {derive_loop_name(r["tag_name"]) for r in rows}
    missing_loops = [l for l in loop_names if l not in loops_with_data]
    if missing_loops:
        raise ValueError(
            f"No data between {start_ts} and {end_ts} for: " + ", ".join(missing_loops)
        )

    pivot: dict = {}
    for r in rows:
        ts = r["ts"]
        pivot.setdefault(ts, {})[r["tag_name"]] = r["value"] if r["value"] is not None else r["text_value"]

    sorted_ts = sorted(pivot.keys())
    wanted_tags_sorted = sorted(wanted_tags)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["TIMESTAMP"] + wanted_tags_sorted)
    for ts in sorted_ts:
        row = pivot[ts]
        ws.append([ts] + [row.get(tag) for tag in wanted_tags_sorted])

    if suffix_map:
        fmt_ws = wb.create_sheet("loop_format")
        roles = ["pv", "op", "sp", "mode"]
        fmt_ws.append(roles)
        candidate_lists = [suffix_map.get(r) or [r.upper()] for r in roles]
        for i in range(max(len(c) for c in candidate_lists)):
            fmt_ws.append([c[i] if i < len(c) else None for c in candidate_lists])

    # The engine's health check hard-requires DIAGNOSTIC_CONFIG and
    # MODE_MAPPING sheets to be present (checks 2.2/2.3) — write them from
    # the app's own defaults so a DB-rebuilt file passes the same checks
    # an Excel upload would. The real threshold values still come from the
    # config dict passed to run_v3(), not from this sheet's contents.
    diag_config_rows, mode_mapping_rows = _load_required_sheet_rows()

    dc_ws = wb.create_sheet("DIAGNOSTIC_CONFIG")
    dc_ws.append(["Parameter", "Value"])
    for param, val in diag_config_rows:
        dc_ws.append([param, val])

    mm_ws = wb.create_sheet("MODE_MAPPING")
    mm_ws.append(["Category", "Value"])
    for cat, val in mode_mapping_rows:
        mm_ws.append([cat, val])

    wb.save(out_path)
    return len(sorted_ts)
