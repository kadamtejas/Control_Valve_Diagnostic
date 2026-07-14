"""
main.py - FastAPI backend for Valve Diagnostics POC (with auth + upload + background jobs)

Run with:
    cd D:\\desktop\\Ingenero\\Valve_Diagnostic_Tool_POC\\dashboard
    python -m uvicorn main:app --reload --port 8001

Then open: http://localhost:8001
"""

import json
import os
import shutil
import subprocess
import sys
import threading
import uuid
from datetime import datetime
from pathlib import Path

from fastapi import (
    Depends, FastAPI, File, HTTPException, Request, UploadFile, status
)
from fastapi.responses import (
    FileResponse, HTMLResponse, JSONResponse, RedirectResponse,
    StreamingResponse,
)
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates

from auth import authenticate_user, create_access_token, get_current_user, register_user
from reader import (
    read_all_loop_names,
    read_dashboard_data,
    read_loop_timeseries,
    read_unit_mapping,
)
import db_store
from dotenv import load_dotenv
from groq import Groq

# ── Paths ─────────────────────────────────────────────────────────────────────
# When running as a PyInstaller .exe, VALVE_BASE_DIR is set by launcher.py
# to the folder containing the .exe. In normal dev mode, use the old logic.
if os.environ.get("VALVE_BASE_DIR"):
    BASE_DIR = Path(os.environ["VALVE_BASE_DIR"])
else:
    BASE_DIR = Path(__file__).parent.parent   # …/Valve_Diagnostic_Tool_POC

DASHBOARD_DIR = Path(__file__).parent
UPLOADS_DIR   = BASE_DIR / "uploads"
UPLOADS_DIR.mkdir(exist_ok=True)

# ── Paths
DEFAULT_CONFIG_PATH = BASE_DIR / "dashboard" / "default_config.json"
USER_CONFIG_DIR = BASE_DIR / "user_configs"
USER_CONFIG_DIR.mkdir(exist_ok=True)

# ── Chatbot (Groq) ────────────────────────────────────────────────────────────
load_dotenv(BASE_DIR / ".env")
KNOWLEDGE_PATH = BASE_DIR / "chatbot_knowledge.md"
GROQ_MODEL = "llama-3.3-70b-versatile"
CHAT_HISTORY_LIMIT = 10
GROQ_TIMEOUT = 20.0  # per-key timeout; worst-case wait = num_keys x GROQ_TIMEOUT
_groq_clients = None


def _load_groq_keys():
    """Collect Groq API keys from env, in priority order, de-duplicated.
    Supports GROQ_API_KEY, GROQ_API_KEY_2, GROQ_API_KEY_3, and a
    comma-separated GROQ_API_KEYS. The first key is tried first."""
    keys = []
    multi = os.environ.get("GROQ_API_KEYS", "")
    if multi.strip():
        keys.extend(k.strip() for k in multi.split(",") if k.strip())
    for name in ("GROQ_API_KEY", "GROQ_API_KEY_2", "GROQ_API_KEY2",
                 "GROQ_API_KEY_3", "GROQ_API_KEY3",
                 "GROQ_API_KEY_4", "GROQ_API_KEY4"):
        v = os.environ.get(name, "").strip()
        if v:
            keys.append(v)
    seen, ordered = set(), []
    for k in keys:
        if k not in seen:
            seen.add(k)
            ordered.append(k)
    return ordered


def _get_groq_clients():
    """Build (once) one Groq client per available key, cached in order."""
    global _groq_clients
    if _groq_clients is None:
        keys = _load_groq_keys()
        if not keys:
            raise RuntimeError("No GROQ_API_KEY found — check the .env file in the POC root.")
        _groq_clients = [Groq(api_key=k, max_retries=0, timeout=GROQ_TIMEOUT) for k in keys]
    return _groq_clients


def _is_rotatable_error(e) -> bool:
    """True only for key-specific failures (rate-limit / auth) where a
    different key might succeed. False for request errors (e.g. 400) that
    would fail identically on every key."""
    status = getattr(e, "status_code", None)
    if status in (429, 401, 403):
        return True
    m = str(e).lower()
    return ("rate limit" in m or "429" in m or "401" in m or "403" in m
            or "invalid api key" in m or "authentication" in m)


def _groq_chat(**kwargs):
    """chat.completions.create with key failover. Tries each key in order;
    on a key-specific failure moves to the next key. Non-rotatable errors
    (and the last key's error) are raised immediately."""
    clients = _get_groq_clients()
    for i, client in enumerate(clients):
        try:
            return client.chat.completions.create(**kwargs)
        except Exception as e:
            if _is_rotatable_error(e) and i < len(clients) - 1:
                continue  # this key is rate-limited / invalid — try the next
            raise


def _groq_chat_stream(**kwargs):
    """Yield reply text chunks with key failover. Failover happens only before
    the first token is emitted; once a key starts streaming we stay on it (a
    stream cannot be cleanly restarted mid-flow). Rotatable errors from Groq
    surface at creation / first chunk, before any emit, so this is safe."""
    clients = _get_groq_clients()
    kwargs["stream"] = True
    for i, client in enumerate(clients):
        emitted = False
        try:
            stream = client.chat.completions.create(**kwargs)
            for chunk in stream:
                delta = ""
                if chunk.choices:
                    delta = chunk.choices[0].delta.content or ""
                if delta:
                    emitted = True
                    yield delta
            return
        except Exception as e:
            if not emitted and _is_rotatable_error(e) and i < len(clients) - 1:
                continue  # nothing sent yet on this key — try the next
            raise


def _load_chatbot_knowledge() -> str:
    """Read the knowledge file fresh on every request so it can be edited live."""
    try:
        return KNOWLEDGE_PATH.read_text(encoding="utf-8")
    except Exception:
        return "(Knowledge file not available.)"


def _knowledge_for_query(user_message: str) -> str:
    """RAG: inject only the knowledge chunks relevant to the question.
    Falls back to the full knowledge file if the index isn't built yet, the
    HF API fails, or nothing scores as relevant — so chat never breaks.
    """
    try:
        from chatbot_rag import retrieve
        hit = retrieve(user_message)
        if hit is not None:        # "" = nothing relevant (send no knowledge); str = chunks
            return hit
    except Exception:
        pass
    return _load_chatbot_knowledge()   # None = retrieval failed -> full-file safety net

# ── In-memory state ───────────────────────────────────────────────────────────
# { email: "results_My_plant_data_1" }          — active results folder per user
user_results: dict[str, str] = {}

# { job_id: { "status": "running"|"done"|"error", "results_dir": "...", "detail": "..." } }
jobs: dict[str, dict] = {}

# { email: { "dest_path": ..., "results_folder": ..., "mode": ... } } — awaiting config confirmation
pending_uploads: dict[str, dict] = {}

# { email: bool } — True means user has explicitly re-run with custom settings this session
user_custom_run: dict[str, bool] = {}

# ── App ───────────────────────────────────────────────────────────────────────
app = FastAPI(title="Valve Diagnostics POC", version="2.0.0")

templates = Jinja2Templates(directory=str(DASHBOARD_DIR / "templates"))

static_dir = DASHBOARD_DIR / "static"
static_dir.mkdir(exist_ok=True)
app.mount("/static", StaticFiles(directory=str(static_dir)), name="static")

SAMPLE_FILE = static_dir / "sample_plant_data.xlsx"


@app.on_event("shutdown")
async def _shutdown_db_pool():
    await db_store.close_pool()


@app.get("/download/manual")
async def download_manual():
    manual = BASE_DIR / "Valve_Diagnostic_Tool_Manual_2.docx"
    if not manual.exists():
        raise HTTPException(status_code=404, detail="Manual not found")
    return FileResponse(
        str(manual),
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        filename="Valve_Diagnostic_Tool_Manual.docx",
    )


@app.get("/download/sample")
async def download_sample():
    if not SAMPLE_FILE.exists():
        raise HTTPException(status_code=404, detail="Sample file not found")
    return FileResponse(
        str(SAMPLE_FILE),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename="sample_plant_data.xlsx",
    )


# ── Helpers ───────────────────────────────────────────────────────────────────

def _results_dir_for(email: str) -> str | None:
    """Return the absolute results folder path for this user, or None."""
    folder = user_results.get(email)
    if folder:
        full = BASE_DIR / folder
        if full.exists():
            return str(full)
    return None


def _run_engine_background(job_id: str, input_path: Path, results_folder: str, mode: str,
                           config: dict = None):
    """
    Runs in a background thread.
    When frozen as .exe: calls run_v3() directly (no subprocess needed — engine is bundled).
    In dev mode: calls valve_diagnostics_v3.py via subprocess like the bat files do.

    `config`, if given, is a flat {parameter: value} dict — the resolved
    per-user config (default or their own saved settings). It's passed
    straight through to the engine; nothing gets written into or read back
    from the uploaded Excel's DIAGNOSTIC_CONFIG sheet anymore.
    """
    output_dir = str(BASE_DIR / results_folder)
    engine_mode = "MANUAL" if mode == "manual" else "AUTO"

    if getattr(sys, 'frozen', False):
        # ── Frozen .exe path: import and call run_v3 directly ─────────────────
        # sys.executable is the .exe itself here, NOT Python, so subprocess
        # won't work. But all engine code is bundled, so we can call it directly.
        try:
            import os as _os
            _os.chdir(str(BASE_DIR))   # engine uses relative paths internally
            # Ensure valve_diagnostics_v2 and column_normaliser are importable
            if str(BASE_DIR) not in sys.path:
                sys.path.insert(0, str(BASE_DIR))
            from diagnostics import run_v3
            exit_code = run_v3(
                str(input_path),
                output_dir,
                mode=engine_mode,
                verbose=True,
                config=config,
            )
            if exit_code == 0:
                jobs[job_id]["status"] = "done"
            elif exit_code == 2:
                jobs[job_id]["status"] = "error"
                jobs[job_id]["detail"] = "Health check failed — check your Excel file format."
            else:
                jobs[job_id]["status"] = "error"
                jobs[job_id]["detail"] = f"Engine exited with code {exit_code}."
        except Exception as exc:
            import traceback
            jobs[job_id]["status"] = "error"
            jobs[job_id]["detail"] = f"{exc}\n{traceback.format_exc()[-600:]}"
    else:
        # ── Dev mode: subprocess call (same as the .bat files) ────────────────
        cmd = [
            sys.executable,
            str(BASE_DIR / "valve_diagnostics_v3.py"),
            "--input", str(input_path),
            "--output-dir", output_dir,
        ]
        if mode == "manual":
            cmd.append("--manual")

        config_json_path = None
        if config is not None:
            os.makedirs(output_dir, exist_ok=True)
            config_json_path = Path(output_dir) / "_run_config.json"
            with open(config_json_path, "w", encoding="utf-8") as f:
                json.dump(config, f, indent=2)
            cmd.extend(["--config-json", str(config_json_path)])

        try:
            result = subprocess.run(
                cmd,
                cwd=str(BASE_DIR),
                capture_output=True,
                encoding='utf-8',
                errors='replace',
            )
            if result.returncode == 0:
                jobs[job_id]["status"] = "done"
            elif result.returncode == 2:
                jobs[job_id]["status"] = "error"
                jobs[job_id]["detail"] = "Health check failed — check your Excel file format."
            else:
                jobs[job_id]["status"] = "error"
                jobs[job_id]["detail"] = (
                    f"Engine exited with code {result.returncode}. "
                    f"{result.stderr[-400:] if result.stderr else ''}"
                )
        except Exception as exc:
            jobs[job_id]["status"] = "error"
            jobs[job_id]["detail"] = str(exc)


# ── Config helpers ────────────────────────────────────────────────────────────

def _get_user_config_path(email: str) -> Path:
    safe = email.replace("@", "_at_").replace(".", "_")
    return USER_CONFIG_DIR / f"{safe}.json"


def _load_config(email: str) -> dict:
    """Load user config, falling back to default."""
    user_path = _get_user_config_path(email)
    if user_path.exists():
        with open(user_path, encoding="utf-8") as f:
            return json.load(f)
    if DEFAULT_CONFIG_PATH.exists():
        with open(DEFAULT_CONFIG_PATH, encoding="utf-8") as f:
            return json.load(f)
    return {"diagnostic_config": [], "diagnostic_selection": [], "mode_mapping": []}


def _is_custom_config(email: str) -> bool:
    """Return True only when the user has explicitly re-run with custom settings."""
    return user_custom_run.get(email, False)


def _save_config(email: str, config: dict):
    # Normalise diagnostic_selection enabled values to int (1/0).
    # Guards against stale "Yes"/"No" strings that survive without being toggled.
    for row in config.get("diagnostic_selection", []):
        v = row.get("enabled")
        if isinstance(v, str):
            row["enabled"] = 0 if v.strip().lower() in ("0", "no", "false", "off", "n", "") else 1
        elif isinstance(v, bool):
            row["enabled"] = 1 if v else 0
    path = _get_user_config_path(email)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(config, f, indent=2)


def _flatten_config(config: dict) -> dict:
    """Convert the JSON-storage config shape (diagnostic_config as a list of
    {parameter, value, description} rows) into the flat {parameter: value}
    dict the engine actually expects. This is the only place that shape
    conversion happens — everything downstream of this deals in flat dicts.
    """
    flat = {}
    for row in config.get("diagnostic_config", []):
        p = row.get("parameter")
        if p is not None:
            flat[p] = row.get("value")
    return flat


# ── Auth endpoints ────────────────────────────────────────────────────────────

@app.get("/login", response_class=HTMLResponse)
async def login_page(request: Request):
    return templates.TemplateResponse(request, "login.html", {"error": None})


@app.post("/login")
async def login_submit(request: Request):
    form     = await request.form()
    email    = form.get("email", "").strip()
    password = form.get("password", "")

    user = authenticate_user(email, password)
    if not user:
        return templates.TemplateResponse(
            request, "login.html",
            {"error": "Invalid email or password."},
            status_code=status.HTTP_401_UNAUTHORIZED,
        )

    token    = create_access_token(user)
    response = RedirectResponse(url="/upload", status_code=302)
    response.set_cookie(
        key="access_token", value=token,
        httponly=True, samesite="lax", max_age=8 * 3600,
    )
    return response


@app.get("/logout")
async def logout():
    response = RedirectResponse(url="/login", status_code=302)
    response.delete_cookie("access_token")
    return response



@app.get("/register", response_class=HTMLResponse)
async def register_page(request: Request):
    return templates.TemplateResponse(request, "register.html", {"error": None, "success": None})


@app.post("/register")
async def register_submit(request: Request):
    form     = await request.form()
    name     = form.get("name", "").strip()
    email    = form.get("email", "").strip()
    password = form.get("password", "")
    confirm  = form.get("confirm_password", "")

    if password != confirm:
        return templates.TemplateResponse(
            request, "register.html",
            {"error": "Passwords do not match.", "success": None},
            status_code=400,
        )
    if len(password) < 6:
        return templates.TemplateResponse(
            request, "register.html",
            {"error": "Password must be at least 6 characters.", "success": None},
            status_code=400,
        )

    success, message = register_user(email, password, name)
    if not success:
        return templates.TemplateResponse(
            request, "register.html",
            {"error": message, "success": None},
            status_code=400,
        )

    return templates.TemplateResponse(
        request, "register.html",
        {"error": None, "success": "Account created! You can now sign in."},
    )


# ── Upload endpoints ──────────────────────────────────────────────────────────

@app.get("/upload", response_class=HTMLResponse)
async def upload_page(request: Request, current_user: dict = Depends(get_current_user)):
    return templates.TemplateResponse(request, "upload.html", {"user": current_user})


@app.post("/upload")
async def upload_file(
    request: Request,
    file: UploadFile = File(...),
    current_user: dict = Depends(get_current_user),
):
    # ── Validate extension
    if not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="Only .xlsx / .xls files are accepted.")

    # ── Read mode and tag suffixes from form
    form = await request.form()
    mode = form.get("mode", "auto").lower()
    if mode not in ("auto", "manual"):
        mode = "auto"

    # Push-to-database is opt-in and admin-only — regular users only ever
    # get the plain Excel flow, no data leaves the file into Postgres.
    # Gated server-side (not just hidden in the UI) since a hidden checkbox
    # alone isn't real access control.
    save_to_db = (
        form.get("save_to_db", "false").lower() == "true"
        and current_user.get("role") == "admin"
    )

    # Tag suffixes — user-defined signal type names e.g. MV, SV, CV, AOUT
    # Fall back to standard PV/SP/OP/MODE if not provided
    raw_suffixes = [
        str(form.get("suffix_pv",   "PV")  or "PV").strip().upper(),
        str(form.get("suffix_sp",   "SP")  or "SP").strip().upper(),
        str(form.get("suffix_op",   "OP")  or "OP").strip().upper(),
        str(form.get("suffix_mode", "MODE") or "MODE").strip().upper(),
    ]
    # Validate: all 4 must be non-empty strings
    tag_suffixes = raw_suffixes if all(raw_suffixes) else ["PV", "SP", "OP", "MODE"]

    # ── Save to user-specific subfolder to prevent cross-user file collisions
    safe_email = current_user["sub"].replace("@", "_at_").replace(".", "_")
    user_upload_dir = BASE_DIR / "uploads" / safe_email
    user_upload_dir.mkdir(parents=True, exist_ok=True)
    original_name = Path(file.filename).name
    dest_path     = user_upload_dir / original_name
    with open(dest_path, "wb") as f:
        shutil.copyfileobj(file.file, f)

    # ── Also persist to Postgres so this data can be re-pulled later from
    #    the "Run from Database" tab without re-uploading. Opt-in, admin-only
    #    (see save_to_db above). Best-effort even then — if Postgres isn't
    #    reachable, the Excel flow must still work regardless.
    if save_to_db:
        try:
            rows_written = await db_store.insert_readings_from_excel(
                str(dest_path), original_name, current_user["sub"]
            )
            print(f"[db_store] {rows_written} readings upserted from {original_name}")
        except Exception as db_exc:
            print(f"[db_store] WARNING — failed to persist to Postgres: {db_exc}")

    # ── Derive results folder name namespaced by user email
    basename       = Path(original_name).stem
    results_folder = f"results_{safe_email}_{basename}" if mode == "auto" else f"results_{safe_email}_{basename}_manual"

    # ── Load default/user config and write it into the Excel
    config = _load_config(current_user["sub"])
    # Clear unit/uom mapping so next GET re-seeds from the new file's Excel
    config.pop("unit_mapping", None)
    config.pop("uom_mapping", None)
    _save_config(current_user["sub"], config)
    flat_config = _flatten_config(config)

    # ── Start engine immediately with defaults — no config page
    job_id = str(uuid.uuid4())
    jobs[job_id] = {
        "status":         "running",
        "results_folder": results_folder,
        "detail":         "",
        "email":          current_user["sub"],
    }
    thread = threading.Thread(
        target=_run_engine_background,
        args=(job_id, dest_path, results_folder, mode, flat_config),
        daemon=True,
    )
    thread.start()

    # ── Also store as pending so the user can later open settings and re-run
    pending_uploads[current_user["sub"]] = {
        "dest_path":      str(dest_path),
        "results_folder": results_folder,
        "mode":           mode,
        "original_name":  original_name,
        "tag_suffixes":   tag_suffixes,
    }

    # ── Fresh upload always resets the custom-run flag (badge will show)
    user_custom_run[current_user["sub"]] = False

    return JSONResponse({"status": "running", "job_id": job_id})


# ── Job polling endpoint ──────────────────────────────────────────────────────

@app.get("/api/job/{job_id}")
async def poll_job(job_id: str, current_user: dict = Depends(get_current_user)):
    job = jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found.")

    # Only the user who started the job can poll it
    if job["email"] != current_user["sub"]:
        raise HTTPException(status_code=403, detail="Not your job.")

    if job["status"] == "done":
        # Register results folder for this user
        user_results[current_user["sub"]] = job["results_folder"]
        return JSONResponse({"status": "done", "redirect": "/"})

    if job["status"] == "error":
        return JSONResponse({"status": "error", "detail": job.get("detail", "Unknown error.")})

    return JSONResponse({"status": "running"})


# ── Database (Postgres) endpoints ───────────────────────────────────────────

@app.get("/api/db/loops")
async def db_list_loops(current_user: dict = Depends(get_current_user)):
    """Loops/tags/date-range currently available in Postgres, for the
    'Run from Database' tab's loop picker."""
    try:
        data = await db_store.list_available_loops()
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"Database unavailable: {exc}")
    return JSONResponse(content=data)


@app.post("/api/db/run")
async def db_run_diagnostics(request: Request, current_user: dict = Depends(get_current_user)):
    """Pull the selected loops/date-range from Postgres, rebuild it as an
    Excel file in the exact Sheet1 layout the engine expects, and run
    diagnostics on it exactly like a normal upload — the engine itself
    never knows the data came from the DB instead of a user-picked file."""
    body = await request.json()
    loops = body.get("loops") or []
    start = body.get("start")
    end   = body.get("end")
    mode  = str(body.get("mode", "auto")).lower()
    suffix_map = body.get("suffix_map") or None
    if mode not in ("auto", "manual"):
        mode = "auto"

    if not loops:
        raise HTTPException(status_code=400, detail="Select at least one loop.")
    if not start or not end:
        raise HTTPException(status_code=400, detail="Select a start and end date/time.")

    safe_email = current_user["sub"].replace("@", "_at_").replace(".", "_")
    user_upload_dir = BASE_DIR / "uploads" / safe_email
    user_upload_dir.mkdir(parents=True, exist_ok=True)

    basename  = f"db_pull_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    dest_path = user_upload_dir / f"{basename}.xlsx"

    try:
        n_rows = await db_store.build_excel_from_db(loops, start, end, str(dest_path), suffix_map=suffix_map)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc))
    except Exception as exc:
        raise HTTPException(status_code=503, detail=f"Database unavailable: {exc}")

    results_folder = f"results_{safe_email}_{basename}" if mode == "auto" else f"results_{safe_email}_{basename}_manual"

    config = _load_config(current_user["sub"])
    config.pop("unit_mapping", None)
    config.pop("uom_mapping", None)
    _save_config(current_user["sub"], config)
    flat_config = _flatten_config(config)

    job_id = str(uuid.uuid4())
    jobs[job_id] = {
        "status":         "running",
        "results_folder": results_folder,
        "detail":         "",
        "email":          current_user["sub"],
    }
    thread = threading.Thread(
        target=_run_engine_background,
        args=(job_id, dest_path, results_folder, mode, flat_config),
        daemon=True,
    )
    thread.start()

    pending_uploads[current_user["sub"]] = {
        "dest_path":      str(dest_path),
        "results_folder": results_folder,
        "mode":           mode,
        "original_name":  f"{basename}.xlsx",
        "tag_suffixes":   ["PV", "SP", "OP", "MODE"],
    }
    user_custom_run[current_user["sub"]] = False

    return JSONResponse({"status": "running", "job_id": job_id, "rows_pulled": n_rows})


# ── Config endpoints ──────────────────────────────────────────────────────────

@app.get("/config", response_class=HTMLResponse)
async def config_page(request: Request, current_user: dict = Depends(get_current_user)):
    # Config page is no longer shown automatically — redirect to dashboard
    return RedirectResponse(url="/", status_code=302)


@app.get("/api/config")
async def get_config(current_user: dict = Depends(get_current_user)):
    config = _load_config(current_user["sub"])
    return JSONResponse(content=config)


@app.post("/api/config/run")
async def save_config_and_run(request: Request, current_user: dict = Depends(get_current_user)):
    pending = pending_uploads.get(current_user["sub"])
    if not pending:
        raise HTTPException(status_code=400, detail="No pending upload. Please upload a file first.")

    config = await request.json()

    # Save user config
    _save_config(current_user["sub"], config)

    # Mark that this user has now run with custom settings
    user_custom_run[current_user["sub"]] = True

    dest_path = Path(pending["dest_path"])
    flat_config = _flatten_config(config)

    # Explicit Settings interaction -- force manual mode so auto-calibration
    # never silently overrides a value the user just set.
    results_folder = pending["results_folder"]
    mode = "manual"
    job_id = str(uuid.uuid4())
    jobs[job_id] = {
        "status":         "running",
        "results_folder": results_folder,
        "detail":         "",
        "email":          current_user["sub"],
    }
    thread = threading.Thread(
        target=_run_engine_background,
        args=(job_id, dest_path, results_folder, mode, flat_config),
        daemon=True,
    )
    thread.start()

    # Clear pending
    pending_uploads.pop(current_user["sub"], None)

    return JSONResponse({"job_id": job_id, "status": "running"})


# ── Re-run with new settings (called from the Settings panel on the dashboard) ──

@app.post("/api/config/rerun")
async def rerun_with_config(request: Request, current_user: dict = Depends(get_current_user)):
    pending = pending_uploads.get(current_user["sub"])

    # ── If pending is missing (e.g. server restarted), recover from results folder on disk
    if not pending:
        rd = _results_dir_for(current_user["sub"])
        if rd:
            folder_name = os.path.basename(rd)
            safe_email  = current_user["sub"].replace("@", "_at_").replace(".", "_")
            is_manual   = folder_name.endswith("_manual")
            stem = _stem_from_results_folder(rd, current_user["sub"])
            mode        = "manual" if is_manual else "auto"
            results_folder = folder_name
            # Look for matching .xlsx in user upload dir
            user_upload_dir = BASE_DIR / "uploads" / safe_email
            input_path = user_upload_dir / f"{stem}.xlsx"
            if not input_path.exists():
                input_path = user_upload_dir / f"{stem}.xls"
            if not input_path.exists():
                raise HTTPException(
                    status_code=400,
                    detail=f"Source file '{stem}.xlsx' not found. Please upload your file again."
                )
            pending = {
                "dest_path":      str(input_path),
                "results_folder": results_folder,
                "mode":           mode,
                "original_name":  input_path.name,
            }
            pending_uploads[current_user["sub"]] = pending
        else:
            raise HTTPException(
                status_code=400,
                detail="No uploaded file found for your session. Please upload a file first."
            )

    config = await request.json()

    # Save user config
    _save_config(current_user["sub"], config)

    dest_path = Path(pending["dest_path"])
    flat_config = _flatten_config(config)

    # Explicit Settings interaction -- force manual mode so auto-calibration
    # never silently overrides a value the user just set.
    results_folder = pending["results_folder"]
    mode = "manual"
    job_id = str(uuid.uuid4())
    jobs[job_id] = {
        "status":         "running",
        "results_folder": results_folder,
        "detail":         "",
        "email":          current_user["sub"],
    }
    thread = threading.Thread(
        target=_run_engine_background,
        args=(job_id, dest_path, results_folder, mode, flat_config),
        daemon=True,
    )
    thread.start()

    return JSONResponse({"job_id": job_id, "status": "running"})


# ── Dashboard ─────────────────────────────────────────────────────────────────

@app.get("/", response_class=HTMLResponse)
async def index(request: Request, current_user: dict = Depends(get_current_user)):
    response = templates.TemplateResponse(request, "index.html", {"user": current_user})
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    return response


# ── API: results (always from THIS user's results folder) ────────────────────

def _get_user_results_dir(current_user: dict) -> str:
    rd = _results_dir_for(current_user["sub"])
    if not rd:
        raise HTTPException(
            status_code=404,
            detail="No results found for your session. Please upload a file first.",
        )
    return rd


def _stem_from_results_folder(rd: str, email: str) -> str:
    """Recover the original uploaded filename's stem from a results folder
    path/name. Results folders are named results_{safe_email}_{stem}, or
    results_{safe_email}_{stem}_manual for manual-mode runs. This strips
    both, leaving just the stem — used both to locate the matching file
    under uploads/{safe_email}/{stem}.xlsx and to build a clean, human-
    readable display name (no internal naming convention shown to the user).
    """
    folder_name = os.path.basename(rd) if rd else ""
    safe_email = email.replace("@", "_at_").replace(".", "_")
    prefix = f"results_{safe_email}_"
    return folder_name.removeprefix(prefix).removesuffix("_manual")


@app.get("/api/latest")
async def get_latest(current_user: dict = Depends(get_current_user)):
    rd = _get_user_results_dir(current_user)
    try:
        data = read_dashboard_data(rd, base_dir=str(BASE_DIR))
        data["is_custom_config"] = _is_custom_config(current_user["sub"])
        data["run_folder_display"] = _stem_from_results_folder(rd, current_user["sub"]) or data.get("run_folder", "")
        return JSONResponse(content=data)
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Error reading results: {e}")


def _fmt_num(v, d=2):
    try:
        return f"{float(v):.{d}f}"
    except (TypeError, ValueError):
        return "-"


def _build_results_summary(email: str, user_message: str = "") -> str:
    """Query-aware per-loop summary. Sends full detail only for loops relevant to the user's question.
    Healthy loops are compressed to one line. Scales cleanly to 30-100 loops.
    """
    # ── Query analysis ────────────────────────────────────────────────────────
    msg_lower = user_message.lower()

    # Topic keywords → which diagnosis types to expand
    TOPIC_MAP = {
        "stiction": ["stiction"],
        "stick": ["stiction"],
        "oscillat": ["aggressive tuning", "external oscillation"],
        "aggress": ["aggressive tuning"],
        "external": ["external oscillation"],
        "sluggish": ["sluggish"],
        "slow": ["sluggish"],
        "saturat": ["saturation"],
        "manual": ["loop in manual"],
        "sensor": ["sensor"],
        "frozen": ["sensor"],
        "data quality": ["data quality"],
        "compression": ["data quality"],
        "propagat": ["_propagation"],
        "tuning": ["aggressive tuning", "sluggish"],
        "worst": ["_worst"],
        "critical": ["_worst"],
        "bad": ["_worst"],
        "overall": ["_plant"],
        "plant": ["_plant"],
        "summary": ["_plant"],
        "health": ["_plant"],
        "mainten": ["_maintenance"],
    }

    # Collect matched topics from the user message
    matched_topics: list = []
    for kw, topics in TOPIC_MAP.items():
        if kw in msg_lower:
            matched_topics.extend(topics)
    matched_topics = list(set(matched_topics))

    rd = _results_dir_for(email)
    if not rd:
        return "The user has NOT run a diagnostic yet. No results are available. Tell them to upload a plant data file on the Upload page first if they ask about results."
    try:
        data = read_dashboard_data(rd, base_dir=str(BASE_DIR))
    except Exception as e:
        return f"Results exist but could not be loaded ({e})."

    lines = []

    # ── Active config thresholds ──────────────────────────────────────────────
    try:
        user_cfg_path = USER_CONFIG_DIR / f"{email.replace('@', '_').replace('.', '_')}.json"
        cfg_path = user_cfg_path if user_cfg_path.exists() else DEFAULT_CONFIG_PATH
        cfg_raw = json.loads(cfg_path.read_text(encoding="utf-8"))
        cfg_params = {row["parameter"]: row["value"] for row in cfg_raw.get("diagnostic_config", [])}
        if cfg_params:
            lines.append("Active diagnostic thresholds (used in this run):")
            threshold_labels = {
                "HARRIS_INDEX_THRESHOLD": "Harris Index threshold (below = poor)",
                "OSCILLATION_REGULARITY_MIN": "Hagglund regularity threshold (above = oscillating)",
                "IAE_PER_HOUR_THRESHOLD": "IAE/hr threshold (above = poor tracking)",
                "OP_ACTIVITY_THRESHOLD": "OP activity threshold (above = high valve movement)",
                "AMP_THRESHOLD": "PV amplitude threshold (above = high oscillation)",
                "SERVICE_FACTOR_MIN_PCT": "Service factor min % (below = loop in manual)",
                "STICT_CONF_HIGH": "Stiction high confidence threshold (%)",
                "STICT_CONF_MED": "Stiction medium confidence threshold (%)",
                "STICTION_S_MIN_PCT": "Stiction S min stickband (% OP)",
                "PROP_CONF_MIN": "Propagation min score to report a link",
            }
            for key, label in threshold_labels.items():
                if key in cfg_params:
                    lines.append(f"  {label}: {cfg_params[key]}")
    except Exception:
        pass

    # ── Plant-level summary ───────────────────────────────────────────────────
    plant = data.get("plant") or {}
    if isinstance(plant, dict):
        phi   = plant.get("plant_health_index")
        total = plant.get("loops_total")
        ana   = plant.get("loops_analysed")
        skip  = plant.get("loops_skipped")
        good  = plant.get("pct_good")
        poor  = plant.get("pct_poor")
        crit  = plant.get("pct_critical")
        ts    = plant.get("run_timestamp", "")
        intv  = plant.get("sample_interval", "")
        dur   = plant.get("duration_hours")
        dcounts = plant.get("diagnosis_counts") or {}
        if phi is not None:
            lines.append(f"Plant Health Index: {_fmt_num(phi, 0)}/100")
        if total:
            # Count health buckets directly from each loop's health value — exact
            # integers that always sum correctly. Deriving from rounded percentages
            # (the old way) drifted and could disagree turn-to-turn.
            n_good = n_poor = n_crit = 0
            for _lp in (data.get("loops") or []):
                try:
                    h = float(_lp.get("health"))
                except (TypeError, ValueError):
                    continue
                if h >= 75:
                    n_good += 1
                elif h >= 50:
                    n_poor += 1
                else:
                    n_crit += 1
            lines.append(f"Loops: total={total}, analysed={ana}, skipped={skip}")
            lines.append(f"Loop health (EXACT counts, do not recompute): "
                         f"healthy(>=75)={n_good}, attention(50-74)={n_poor}, "
                         f"critical(<50)={n_crit}")
            lines.append(f"Loops in trouble (attention + critical) = {n_poor + n_crit} "
                         f"of {total}. Use THIS number for 'how many loops are in trouble'; "
                         f"do NOT count from fault types below.")
        if ts:
            lines.append(f"Run timestamp: {ts}")
        if dur:
            lines.append(f"Data duration: {_fmt_num(dur, 1)} hours, sample interval: {intv}")
        if dcounts:
            dc_str = ", ".join(f"{k}: {v}" for k, v in dcounts.items())
            lines.append(f"Fault types present (these are fault-type tallies, NOT loop "
                         f"counts — for counting loops use the EXACT counts above): {dc_str}")

    # ── Per-loop details (query-aware) ───────────────────────────────────────
    loops = data.get("loops") or []
    all_loop_names = [str(lp.get('loop', '')).lower() for lp in loops]

    # Detect specific loop names mentioned in the query
    mentioned_loops = [
        lp.get('loop') for lp in loops
        if lp.get('loop') and str(lp['loop']).lower() in msg_lower
    ]

    # ── Results-relevance gate ───────────────────────────────────────────────
    # If the question names no loop and matches no results topic (e.g. "hi", or
    # a pure concept question), send NO loop data and NO plant numbers — just a
    # neutral note that a run exists. This keeps greetings clean (no numbers in
    # context = nothing to recite) and cuts almost all per-message tokens.
    if not matched_topics and not mentioned_loops:
        note = ("A completed diagnostic run is available for this user. Do NOT state "
                "plant health numbers or loop results unless the user actually asks; "
                "if they only greet, greet back briefly and offer to help.")
        try:
            (Path(rd) / "chatbot_debug_summary.txt").write_text(note, encoding="utf-8")
        except Exception:
            pass
        return note

    def _should_expand(lp: dict, matched_topics: list, mentioned_loops: list) -> bool:
        """Decide if this loop gets full detail or one-liner."""
        sev = (lp.get('severity') or '').upper()
        diag = (lp.get('diagnosis') or '').lower()
        loop_name = str(lp.get('loop', ''))
        is_healthy = sev in ('OK', 'GOOD', 'HEALTHY', '-') or diag == 'healthy'

        if loop_name in mentioned_loops:
            return True
        if mentioned_loops:
            return False
        if not matched_topics:
            return not is_healthy
        if "_worst" in matched_topics:
            sorted_loops = sorted(loops, key=lambda x: x.get('health') or 100)
            worst_names = [l.get('loop') for l in sorted_loops[:5]]
            return loop_name in worst_names
        if "_plant" in matched_topics or "_maintenance" in matched_topics:
            return not is_healthy
        if "_propagation" in matched_topics:
            return False
        for topic in matched_topics:
            if topic in diag:
                return True
        return False

    def _loop_full_detail(lp: dict) -> list:
        """Full detail lines for a loop."""
        sev  = lp.get('severity') or '-'
        diag = lp.get('diagnosis') or '-'
        parts = [
            f"{lp.get('loop', '?')}:",
            f"health={_fmt_num(lp.get('health'), 0)}",
            f"severity={sev}",
            f"diagnosis={diag}",
            f"confidence={_fmt_num(lp.get('confidence'), 0)}%",
            f"stiction={lp.get('stiction_label') or '-'}",
            f"service%={_fmt_num(lp.get('service_factor'), 0)}",
            f"IAE/hr={_fmt_num(lp.get('iae_per_hour'))}",
            f"IAEnorm%={_fmt_num(lp.get('iae_per_hour_norm'))}",
            f"harris={_fmt_num(lp.get('harris_index'))}",
            f"PVamp={_fmt_num(lp.get('pv_amplitude'))}",
            f"OPact={_fmt_num(lp.get('op_activity'))}",
            f"hagglund={_fmt_num(lp.get('hagglund_regularity'))}",
            f"dom_period={lp.get('dominant_period') or '-'}",
            f"dataQ={lp.get('data_quality_status') or '-'}",
        ]
        if lp.get('issues'):
            parts.append(f"dataQ_issues={lp['issues']}")
        stiction = lp.get('stiction') or {}
        if stiction:
            parts.append(
                f"stiction_methods(heuristic={_fmt_num(stiction.get('heuristic'))}"
                f"/horch={_fmt_num(stiction.get('horch_cc'))}"
                f"/yamashita={_fmt_num(stiction.get('yamashita'))}"
                f"/bicoh={_fmt_num(stiction.get('bicoherence'))}"
                f" S%={_fmt_num(stiction.get('estimated_s'))})"
            )
        dq = lp.get('data_quality') or {}
        if dq and lp.get('data_quality_status', '').upper() != 'OK':
            parts.append(
                f"dq_detail(missing%={_fmt_num(dq.get('pct_missing'))}"
                f" outliers={dq.get('outliers') or 0}"
                f" frozen={dq.get('frozen') or '-'}"
                f" quantised={dq.get('quantised') or '-'})"
            )
        if lp.get('recommended_action'):
            parts.append(f"action={lp['recommended_action']}")
        result = [" ".join(parts)]
        if lp.get('rationale'):
            result.append(f"  rationale: {lp['rationale'][:300]}")
        return result

    EXPAND_CAP = 8   # max loops shown in full on a topic/generic question
    lines.append(f"\nLoops analysed: {len(loops)}")
    n_expanded = 0
    n_capped = 0
    for lp in loops:
        loop_name = str(lp.get('loop', ''))
        want = _should_expand(lp, matched_topics, mentioned_loops)
        is_named = loop_name in mentioned_loops
        if want and (is_named or n_expanded < EXPAND_CAP):
            lines.extend(_loop_full_detail(lp))
            n_expanded += 1
        else:
            if want:
                n_capped += 1   # wanted full detail but hit the cap
            lines.append(
                f"{lp.get('loop','?')}: {lp.get('diagnosis','-')} "
                f"health={_fmt_num(lp.get('health'),0)} [{lp.get('severity','-')}]"
            )
    if n_capped:
        lines.append(f"(+{n_capped} more flagged loops collapsed to one-liners — ask about "
                     f"a specific loop or category for full detail.)")

    # ── Maintenance actions ───────────────────────────────────────────────────
    maintenance = data.get("maintenance") or []
    if maintenance:
        lines.append("\nMaintenance priority (loops needing physical attention):")
        for m in maintenance[:5]:  # top 5 only
            lines.append(
                f"  {m.get('loop','?')}: {m.get('diagnosis','-')} "
                f"severity={m.get('severity','-')} "
                f"action={m.get('recommended_action','-')}"
            )

    # ── Propagation ───────────────────────────────────────────────────────────
    prop = data.get("propagation") or []
    if prop:
        lines.append("\nCross-loop propagation (oscillation spreading between loops):")
        for p in prop[:10]:
            if isinstance(p, dict):
                lines.append(
                    f"  {p.get('source','?')} -> {p.get('target','?')}: "
                    f"score={_fmt_num(p.get('combined_score'),1)} "
                    f"lag={p.get('lag_time','-')}min "
                    f"corr={_fmt_num(p.get('cross_correlation'),3)} "
                    f"coherence={_fmt_num(p.get('coherence_score'),1)}"
                )

    summary = "\n".join(lines)
    try:
        debug_path = Path(rd) / "chatbot_debug_summary.txt"
        debug_path.write_text(summary, encoding="utf-8")
    except Exception:
        pass
    return summary


CHATBOT_ROLE = (
    "You are the built-in assistant of the Control Valve Diagnostic Tool by Ingenero Technologies. "
    "You help plant operators, control engineers, and plant managers understand the tool, their "
    "diagnostic results, and general process control concepts (PID tuning, stiction, oscillation, etc.). "
    "Use the KNOWLEDGE section for questions about the tool and the LATEST RESULTS section for "
    "questions about the user's own loops. "
    "Reply like a knowledgeable colleague, not a textbook. Be direct and short — 1 to 4 sentences max. "
    "Plain English first, technical terms only if needed. No bullet points unless listing 3+ items. "
    "Exception: when the user asks for a formula, equation, or how a metric is calculated, quote the "
    "exact equation(s) from the KNOWLEDGE section verbatim instead of summarizing — the length limit "
    "does not apply to the equation lines themselves. "
    "Format equations with Markdown: put a standalone equation inside a triple-backtick code block, and "
    "present a multi-step derivation as a numbered list with each step's formula wrapped in single "
    "backticks. Do not use code blocks or backticks in normal non-formula replies. "
    "Never start with 'Great question!', 'Certainly!' or similar. "
    "If unsure, say so in one line — don't guess. Never invent loop names or values not in the results. "
    "If asked something unrelated to the tool or process control, decline in one sentence and steer back."
)


@app.post("/api/chat")
async def api_chat(payload: dict, current_user: dict = Depends(get_current_user)):
    messages = payload.get("messages") or []
    if not isinstance(messages, list) or not messages:
        raise HTTPException(status_code=400, detail="messages list required")

    # keep only the last N well-formed user/assistant turns
    history = [
        {"role": m.get("role"), "content": str(m.get("content", ""))[:4000]}
        for m in messages
        if isinstance(m, dict) and m.get("role") in ("user", "assistant") and m.get("content")
    ][-CHAT_HISTORY_LIMIT:]
    if not history:
        raise HTTPException(status_code=400, detail="no valid messages")

    # Extract the latest user message for query-aware context
    latest_user_msg = next(
        (m["content"] for m in reversed(history) if m["role"] == "user"), ""
    )

    system_prompt = (
        CHATBOT_ROLE
        + "\n\n===== KNOWLEDGE =====\n"
        + _knowledge_for_query(latest_user_msg)
        + "\n\n===== LATEST RESULTS (this user) =====\n"
        + _build_results_summary(current_user["sub"], latest_user_msg)
    )

    try:
        gen = _groq_chat_stream(
            model=GROQ_MODEL,
            messages=[{"role": "system", "content": system_prompt}] + history,
            max_tokens=700,
            temperature=0.3,
        )
        # Pull the first chunk eagerly so auth / rate-limit errors surface as a
        # proper HTTP error instead of a broken 200 stream.
        try:
            first_chunk = next(gen)
        except StopIteration:
            first_chunk = ""

        def event_stream():
            if first_chunk:
                yield first_chunk
            yield from gen

        return StreamingResponse(event_stream(), media_type="text/plain; charset=utf-8")
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        msg = str(e)
        if "rate limit" in msg.lower() or "429" in msg:
            detail = "The chatbot is rate-limited right now. Please wait a minute and try again."
        else:
            detail = f"Chatbot service unavailable: {msg[:200]}"
        raise HTTPException(status_code=503, detail=detail)


@app.get("/api/chat/runid")
async def api_chat_runid(current_user: dict = Depends(get_current_user)):
    """Lightweight id of the user's current diagnostic run — used by the chat
    widget to reset the conversation when a new run appears."""
    rd = _results_dir_for(current_user["sub"])
    if not rd:
        return JSONResponse(content={"run_id": "none"})
    p = Path(rd)
    f = p / "Loop_diagnostics_v2.xlsx"
    try:
        mtime = int((f if f.exists() else p).stat().st_mtime)
    except OSError:
        mtime = 0
    return JSONResponse(content={"run_id": f"{p.name}|{mtime}"})


@app.get("/api/plots")
async def get_plots(current_user: dict = Depends(get_current_user)):
    rd = _get_user_results_dir(current_user)
    rd_path = Path(rd)
    plots = []
    for fname in ["diagnostic_heatmap.png", "plant_dashboard.png"]:
        if (rd_path / fname).exists():
            plots.append({"name": fname, "type": "overview", "url": f"/plot-image/{fname}"})
    plots_dir = rd_path / "plots"
    if plots_dir.exists():
        for f in sorted(plots_dir.iterdir()):
            if f.suffix.lower() == ".png":
                plots.append({"name": f.name, "type": "loop", "url": f"/plot-image/plots/{f.name}"})
    return JSONResponse(content={"plots": plots, "run_folder": rd_path.name})


@app.api_route("/plot-image/{file_path:path}", methods=["GET", "HEAD"])
async def serve_plot(file_path: str, request: Request, current_user: dict = Depends(get_current_user)):
    rd = _get_user_results_dir(current_user)
    full_path = Path(rd) / file_path
    if not full_path.exists():
        raise HTTPException(status_code=404, detail=f"Image not found: {file_path}")
    if request.method == "HEAD":
        return JSONResponse(status_code=200, content={})
    return FileResponse(str(full_path), media_type="image/png")


@app.get("/api/timeseries/{loop_name:path}")
async def get_timeseries(loop_name: str, current_user: dict = Depends(get_current_user)):
    rd = _get_user_results_dir(current_user)
    try:
        data = read_loop_timeseries(rd, loop_name, base_dir=str(BASE_DIR))
        return JSONResponse(content=data)
    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except ValueError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error reading time series: {e}")


@app.get("/api/loops")
async def get_loop_names(current_user: dict = Depends(get_current_user)):
    rd = _get_user_results_dir(current_user)
    loops = read_all_loop_names(rd, base_dir=str(BASE_DIR))
    return JSONResponse(content={"loops": loops})


@app.get("/api/unit-mapping")
async def get_unit_mapping(current_user: dict = Depends(get_current_user)):
    rd = _get_user_results_dir(current_user)
    config = _load_config(current_user["sub"])
    needs_save = False

    # Resolve the actual uploaded input file for this user's current results
    # folder. NOTE: this used to be guessed inside read_unit_mapping() by
    # stripping a generic "results_" prefix from the folder name and looking
    # in BASE_DIR directly — but results folders are actually named
    # results_{safe_email}_{stem}, and uploads live under uploads/{safe_email}/,
    # not BASE_DIR. That mismatch meant the file was never found, so every
    # user's unit map silently came back empty and every loop fell back to
    # "Unknown". Mirrors the same resolution /api/unit-mapping/save uses.
    safe_email = current_user["sub"].replace("@", "_at_").replace(".", "_")
    stem = _stem_from_results_folder(rd, current_user["sub"])
    input_path = str(BASE_DIR / "uploads" / safe_email / f"{stem}.xlsx") if stem else ""

    # If unit_mapping not in JSON yet, seed from Excel
    if not config.get("unit_mapping"):
        excel_data = read_unit_mapping(input_path)
        config["unit_mapping"] = excel_data.get("unit_map", {})
        config["uom_mapping"]  = excel_data.get("uom_map", {})
        needs_save = True

    # If uom_mapping not in JSON yet, seed from Excel
    if "uom_mapping" not in config:
        excel_data = read_unit_mapping(input_path)
        config["uom_mapping"] = excel_data.get("uom_map", {})
        needs_save = True

    if needs_save:
        _save_config(current_user["sub"], config)

    return JSONResponse(content={
        "unit_map": config.get("unit_mapping", {}),
        "units":    sorted(set(config.get("unit_mapping", {}).values())),
        "uom_map":  config.get("uom_mapping", {}),
    })


@app.post("/api/unit-mapping/save")
async def save_unit_mapping(request: Request, current_user: dict = Depends(get_current_user)):
    """
    Persist the unit mapping sent from the UI.
    Payload: { "unit_mapping": { "loopTag": "unitName", ... } }
    The mapping is written back into the UNIT_MAPPING sheet of the original
    input Excel file so it survives server restarts.
    """
    body = await request.json()
    unit_mapping: dict = body.get("unit_mapping", {})
    uom_mapping: dict = body.get("uom_mapping", {})
    if not isinstance(unit_mapping, dict):
        raise HTTPException(status_code=400, detail="unit_mapping must be an object.")

    # ── Find the input Excel for this user's current results folder ──────────
    rd = _get_user_results_dir(current_user)
    safe_email  = current_user["sub"].replace("@", "_at_").replace(".", "_")
    stem = _stem_from_results_folder(rd, current_user["sub"])
    user_upload_dir = BASE_DIR / "uploads" / safe_email
    input_path = user_upload_dir / f"{stem}.xlsx"

    if not input_path.exists():
        raise HTTPException(
            status_code=404,
            detail=f"Source file '{stem}.xlsx' not found. Please upload the file again.",
        )

    # ── Write UNIT_MAPPING sheet ─────────────────────────────────────────────
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(input_path))

        # Create sheet if missing
        if "UNIT_MAPPING" not in wb.sheetnames:
            ws = wb.create_sheet("UNIT_MAPPING")
            ws.cell(1, 1, "Tag")
            ws.cell(1, 2, "Unit")
        else:
            ws = wb["UNIT_MAPPING"]
            # Clear existing data rows
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.value = None

        for i, (tag, unit) in enumerate(unit_mapping.items(), start=2):
            ws.cell(i, 1, str(tag).strip())
            ws.cell(i, 2, str(unit).strip())
            ws.cell(i, 3, str(uom_mapping.get(tag, '')).strip())

        wb.save(str(input_path))
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"Failed to write unit mapping: {exc}")

    return JSONResponse(content={"status": "ok", "saved": len(unit_mapping)})


# ── Report download ───────────────────────────────────────────────────────────

@app.get("/api/report/download")
async def download_report(current_user: dict = Depends(get_current_user)):  # noqa: C901
    """
    Generates and returns a comprehensive PDF report combining:
      - Plant Health Summary  (Loop_diagnostics_v2.xlsx → Plant_Dashboard sheet)
      - Per-loop diagnosis with full explanations (Summary sheet)
      - Outlier handling section (outlier_handling_report.txt)
    """
    import io, re, math
    from datetime import datetime
    from fastapi.responses import StreamingResponse

    rd = _get_user_results_dir(current_user)
    if not rd:
        raise HTTPException(status_code=404, detail="No results found. Run diagnostics first.")

    rd_path = Path(rd)
    xl_path  = rd_path / "Loop_diagnostics_v2.xlsx"
    out_path = rd_path / "outlier_handling_report.txt"

    if not xl_path.exists():
        raise HTTPException(status_code=404, detail="Loop_diagnostics_v2.xlsx not found in results folder.")

    # ── Imports ───────────────────────────────────────────────────────────────
    import pandas as pd
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
        PageBreak, HRFlowable, Flowable,
    )
    from reportlab.graphics.shapes import (
        Drawing, Circle, Rect, String, Line, Wedge,
    )
    from reportlab.graphics import renderPDF

    # ── Helper: round a value safely ─────────────────────────────────────────
    def fmt(v, decimals=1):
        try:
            f = float(str(v))
            return f"{f:.{decimals}f}"
        except:
            return str(v) if str(v) not in ("nan", "None", "") else "—"

    # ── Read data ─────────────────────────────────────────────────────────────
    dash_df = pd.read_excel(xl_path, sheet_name="Plant_Dashboard", header=None)
    kpi, mode = {}, "kpi"
    for _, row in dash_df.iterrows():
        vals = [str(v).strip() if str(v) != "nan" else "" for v in row.values]
        if vals[0] == "Loop" and vals[1] == "Health":
            mode = "loops"; continue
        if mode == "kpi" and vals[0] and vals[1]:
            kpi[vals[0]] = vals[1]

    sum_df = pd.read_excel(xl_path, sheet_name="Summary")

    outlier_text = ""
    if out_path.exists():
        raw = out_path.read_text(encoding="utf-8", errors="ignore")
        outlier_text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f-\x9f═─]', '', raw).strip()

    # ── Extract KPIs ──────────────────────────────────────────────────────────
    phi       = kpi.get("Plant Health Index (0-100)", "0")
    total     = kpi.get("Loops total", "0")
    analysed  = kpi.get("Loops analysed", "0")
    skipped   = kpi.get("Loops skipped", "0")
    good_pct  = kpi.get("% Good (>=75)", "0")
    poor_pct  = kpi.get("% Poor (50–74)", "0")
    crit_pct  = kpi.get("% Critical (<50)", "0")
    duration  = kpi.get("Duration (hours)", "—")
    interval  = kpi.get("Sample interval", "—")
    run_ts    = kpi.get("Run timestamp", "—")

    try:   phi_f = float(phi)
    except: phi_f = 0.0

    # ── Styles ────────────────────────────────────────────────────────────────
    NAVY  = colors.HexColor("#0f2744")
    BLUE  = colors.HexColor("#2e86de")
    OK    = colors.HexColor("#1a9e5a")
    WARN  = colors.HexColor("#d97706")
    CRIT  = colors.HexColor("#dc2626")
    LGREY = colors.HexColor("#f0f4f8")
    MGREY = colors.HexColor("#d8e2ec")
    DGREY = colors.HexColor("#6b8299")
    WHITE = colors.white
    OK_BG   = colors.HexColor("#e6f7ef")
    WARN_BG = colors.HexColor("#fef3e2")
    CRIT_BG = colors.HexColor("#fdf0ef")

    phi_color = OK if phi_f >= 75 else (WARN if phi_f >= 50 else CRIT)
    phi_bg    = OK_BG if phi_f >= 75 else (WARN_BG if phi_f >= 50 else CRIT_BG)

    styles = getSampleStyleSheet()
    def sty(name, **kw):
        s = styles[name].clone(name + str(id(kw)))
        for k, v in kw.items(): setattr(s, k, v)
        return s

    H1   = sty("Heading1", fontSize=20, textColor=NAVY,  spaceAfter=4,  spaceBefore=0,  fontName="Helvetica-Bold")
    H2   = sty("Heading2", fontSize=13, textColor=NAVY,  spaceAfter=6,  spaceBefore=16, fontName="Helvetica-Bold")
    H3   = sty("Heading3", fontSize=10, textColor=NAVY,  spaceAfter=3,  spaceBefore=0,  fontName="Helvetica-Bold")
    BODY = sty("Normal",   fontSize=9,  textColor=colors.HexColor("#3a5068"), spaceAfter=4, leading=14)
    META = sty("Normal",   fontSize=8,  textColor=DGREY, spaceAfter=2)
    CELL = sty("Normal",   fontSize=8,  textColor=colors.HexColor("#3a5068"), leading=12, spaceAfter=0)
    MONO_CELL = sty("Normal", fontSize=8, textColor=NAVY, fontName="Courier", leading=12, spaceAfter=0)

    def hr(color=MGREY, thick=0.5):
        return HRFlowable(width="100%", thickness=thick, color=color, spaceAfter=6, spaceBefore=6)

    def to_hex(c):
        """Convert a reportlab Color object to a 6-char hex string."""
        return '%02x%02x%02x' % (int(c.red*255), int(c.green*255), int(c.blue*255))

    def sev_color(sev):
        s = str(sev).upper()
        if s in ("FAIL","CRITICAL"): return CRIT
        if s in ("WARN","WARNING"):  return WARN
        return OK

    def sev_bg(sev):
        s = str(sev).upper()
        if s in ("FAIL","CRITICAL"): return CRIT_BG
        if s in ("WARN","WARNING"):  return WARN_BG
        return OK_BG

    def sev_hdr_bg(sev): pass
    def sev_hdr_accent(sev): pass
    def sev_act_bg(sev): pass

    # ── Custom Flowable: PHI Gauge ────────────────────────────────────────────
    class PhiGauge(Flowable):
        """Semicircle gauge for the Plant Health Index."""
        def __init__(self, phi_val, color, width=200, height=120):
            super().__init__()
            self.phi_val = phi_val
            self.color   = color
            self.width   = width
            self.height  = height

        def draw(self):
            c = self.canv
            cx = self.width / 2
            cy = 18          # baseline of the semicircle
            r_outer = 70
            r_inner = 48

            # Background arc (grey) — full 180°
            c.setStrokeColor(MGREY)
            c.setFillColor(LGREY)
            c.setLineWidth(0)
            # Draw as a filled annular sector using path
            def arc_path(cx, cy, r, start_deg, end_deg, steps=60):
                pts = []
                for i in range(steps + 1):
                    a = math.radians(start_deg + (end_deg - start_deg) * i / steps)
                    pts.append((cx + r * math.cos(a), cy + r * math.sin(a)))
                return pts

            # Full background ring
            outer_pts = arc_path(cx, cy, r_outer, 0, 180)
            inner_pts = arc_path(cx, cy, r_inner, 180, 0)
            p = c.beginPath()
            p.moveTo(*outer_pts[0])
            for pt in outer_pts[1:]: p.lineTo(*pt)
            for pt in inner_pts: p.lineTo(*pt)
            p.close()
            c.setFillColor(LGREY)
            c.drawPath(p, fill=1, stroke=0)

            # Colored fill arc — from left (180°) sweeping right by phi%
            fill_end = self.phi_val / 100.0 * 180.0
            if fill_end > 1:
                outer_pts2 = arc_path(cx, cy, r_outer, 180 - fill_end, 180)
                inner_pts2 = arc_path(cx, cy, r_inner, 180, 180 - fill_end)
                p2 = c.beginPath()
                p2.moveTo(*outer_pts2[0])
                for pt in outer_pts2[1:]: p2.lineTo(*pt)
                for pt in inner_pts2: p2.lineTo(*pt)
                p2.close()
                c.setFillColor(self.color)
                c.drawPath(p2, fill=1, stroke=0)

            # Centre white circle (donut hole look)
            c.setFillColor(WHITE)
            c.circle(cx, cy, r_inner - 2, fill=1, stroke=0)

            # PHI value text in centre
            c.setFillColor(self.color)
            c.setFont("Helvetica-Bold", 22)
            phi_str = f"{self.phi_val:.0f}"
            c.drawCentredString(cx, cy + 18, phi_str)
            c.setFont("Helvetica", 9)
            c.setFillColor(DGREY)
            c.drawCentredString(cx, cy + 6, "out of 100")

            # Scale labels
            c.setFont("Helvetica", 7)
            c.setFillColor(DGREY)
            c.drawCentredString(cx - r_outer - 4, cy - 4, "0")
            c.drawCentredString(cx, cy + r_outer + 8, "50")
            c.drawCentredString(cx + r_outer + 4, cy - 4, "100")

        def wrap(self, aw, ah):
            return self.width, self.height

    # ── Flowable: coloured stat tile (used in a Table cell) ───────────────────
    def stat_tile(label, value, sub, bg, val_color):
        """Returns a tiny table acting as a coloured tile."""
        tile_data = [[
            Paragraph(f'<font color="#{to_hex(val_color)}"><b>{value}</b></font>',
                      sty("Normal", fontSize=16, fontName="Helvetica-Bold", spaceAfter=0, leading=18)),
        ],[
            Paragraph(f'<b>{label}</b>',
                      sty("Normal", fontSize=8, textColor=NAVY, spaceAfter=0, leading=10)),
        ],[
            Paragraph(sub, sty("Normal", fontSize=7, textColor=DGREY, spaceAfter=0, leading=9)),
        ]]
        t = Table(tile_data, colWidths=[3.8*cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), bg),
            ("PADDING",    (0,0), (-1,-1), 8),
            ("ROUNDEDCORNERS", [6]),
            ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
            ("ALIGN",      (0,0), (-1,-1), "CENTER"),
        ]))
        return t

    # ── sorted loops (used in sections 2 and 3) ───────────────────────────────
    sorted_loops = sorted(
        sum_df.itertuples(),
        key=lambda r: float(str(r._4)) if str(r._4) not in ("nan","") else 999
    )

    # ── Build story ───────────────────────────────────────────────────────────
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        leftMargin=2*cm, rightMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm,
        title="Valve Diagnostics Report"
    )
    story = []

    # ══ COVER HEADER ══════════════════════════════════════════════════════════
    # Navy banner
    banner_data = [[
        Paragraph("<font color='#ffffff'><b>Valve Diagnostics</b></font>",
                  sty("Normal", fontSize=18, fontName="Helvetica-Bold", spaceAfter=0)),
        Paragraph(f"<font color='#7fd3ff'>Comprehensive Diagnostic Report</font><br/>"
                  f"<font color='#6a92b5' size='8'>Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}</font>",
                  sty("Normal", fontSize=11, spaceAfter=0, leading=16)),
    ]]
    banner_tbl = Table(banner_data, colWidths=[7*cm, 10*cm])
    banner_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), NAVY),
        ("PADDING",    (0,0), (-1,-1), 14),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
    ]))
    story.append(banner_tbl)
    story.append(Spacer(1, 0.5*cm))

    # ══ SECTION 1: PLANT HEALTH SUMMARY ══════════════════════════════════════
    story.append(Paragraph("1. Plant Health Summary", H2))

    # Gauge + stat tiles side by side
    try:
        n_good = int(float(total) * float(good_pct) / 100 + 0.5) if good_pct not in ("—","") else "—"
        n_poor = int(float(total) * float(poor_pct) / 100 + 0.5) if poor_pct not in ("—","") else "—"
        n_crit = int(float(total) * float(crit_pct) / 100 + 0.5) if crit_pct not in ("—","") else "—"
    except:
        n_good = n_poor = n_crit = "—"

    gauge = PhiGauge(phi_f, phi_color, width=210, height=130)

    tiles_data = [
        [stat_tile("CRITICAL", str(n_crit),  f"{fmt(crit_pct,0)}% of loops", CRIT_BG, CRIT)],
        [stat_tile("ATTENTION", str(n_poor), f"{fmt(poor_pct,0)}% of loops", WARN_BG, WARN)],
        [stat_tile("HEALTHY",  str(n_good),  f"{fmt(good_pct,0)}% of loops", OK_BG,   OK)],
    ]
    tiles_tbl = Table(tiles_data, colWidths=[4.2*cm], rowHeights=[3.8*cm]*3)
    tiles_tbl.setStyle(TableStyle([
        ("PADDING",  (0,0), (-1,-1), 4),
        ("VALIGN",   (0,0), (-1,-1), "MIDDLE"),
    ]))

    # Info strip below gauge
    info_data = [
        [Paragraph("<b>Total Loops</b>", sty("Normal", fontSize=8, textColor=DGREY, spaceAfter=0)),
         Paragraph("<b>Analysed</b>",    sty("Normal", fontSize=8, textColor=DGREY, spaceAfter=0)),
         Paragraph("<b>Duration</b>",    sty("Normal", fontSize=8, textColor=DGREY, spaceAfter=0)),
         Paragraph("<b>Interval</b>",    sty("Normal", fontSize=8, textColor=DGREY, spaceAfter=0)),
         Paragraph("<b>Run At</b>",      sty("Normal", fontSize=8, textColor=DGREY, spaceAfter=0))],
        [Paragraph(f"<b>{total}</b>",    sty("Normal", fontSize=13, textColor=NAVY, fontName="Helvetica-Bold", spaceAfter=0)),
         Paragraph(f"<b>{analysed}</b>", sty("Normal", fontSize=13, textColor=NAVY, fontName="Helvetica-Bold", spaceAfter=0)),
         Paragraph(f"<b>{fmt(duration,1)} hrs</b>", sty("Normal", fontSize=11, textColor=NAVY, fontName="Helvetica-Bold", spaceAfter=0)),
         Paragraph(f"<b>{interval}</b>", sty("Normal", fontSize=11, textColor=NAVY, fontName="Helvetica-Bold", spaceAfter=0)),
         Paragraph(f"{str(run_ts)[:16]}",sty("Normal", fontSize=8,  textColor=NAVY, spaceAfter=0))],
    ]
    info_tbl = Table(info_data, colWidths=[2.5*cm]*5)
    info_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), LGREY),
        ("GRID",       (0,0), (-1,-1), 0.4, MGREY),
        ("PADDING",    (0,0), (-1,-1), 7),
        ("ALIGN",      (0,0), (-1,-1), "CENTER"),
        ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
    ]))

    # Outer layout: gauge | tiles
    outer_data = [[gauge, tiles_tbl]]
    outer_tbl  = Table(outer_data, colWidths=[11*cm, 6*cm])
    outer_tbl.setStyle(TableStyle([
        ("VALIGN",  (0,0), (-1,-1), "MIDDLE"),
        ("PADDING", (0,0), (-1,-1), 0),
    ]))
    story.append(outer_tbl)
    story.append(Spacer(1, 0.3*cm))
    story.append(info_tbl)
    story.append(Spacer(1, 0.3*cm))

    # Interpretation banner
    if phi_f >= 75:
        interp = f"✔  Plant is in <b>good health</b> (PHI = {phi_f:.1f}/100). Most loops are performing within acceptable limits."
    elif phi_f >= 50:
        interp = f"⚠  Plant is in <b>moderate condition</b> (PHI = {phi_f:.1f}/100). Several loops need attention — review the critical loops below."
    else:
        interp = f"✘  Plant is in <b>poor condition</b> (PHI = {phi_f:.1f}/100). Immediate attention required. Prioritise the critical faults listed below."

    interp_tbl = Table([[Paragraph(interp, sty("Normal", fontSize=9, textColor=phi_color, fontName="Helvetica-Bold", spaceAfter=0))]],
                       colWidths=[17*cm])
    interp_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), phi_bg),
        ("PADDING",    (0,0), (-1,-1), 10),
        ("ROUNDEDCORNERS", [6]),
    ]))
    story.append(interp_tbl)

    # ── Diagnosis breakdown bar ───────────────────────────────────────────────
    story.append(Spacer(1, 0.4*cm))
    diag_counts = sum_df["Diagnosis"].value_counts()
    if not diag_counts.empty:
        diag_header = [Paragraph("<b>Fault Type</b>", sty("Normal", fontSize=8, textColor=WHITE, spaceAfter=0)),
                       Paragraph("<b>Count</b>",       sty("Normal", fontSize=8, textColor=WHITE, spaceAfter=0)),
                       Paragraph("<b>% of Loops</b>",  sty("Normal", fontSize=8, textColor=WHITE, spaceAfter=0))]
        diag_rows = [diag_header]
        total_l = int(float(total)) if str(total).replace(".","").isdigit() else 1
        for fault, cnt in diag_counts.items():
            pct = cnt / total_l * 100
            diag_rows.append([
                Paragraph(str(fault), CELL),
                Paragraph(str(cnt),   sty("Normal", fontSize=8, textColor=NAVY, fontName="Helvetica-Bold", spaceAfter=0)),
                Paragraph(f"{pct:.0f}%", CELL),
            ])
        diag_tbl = Table(diag_rows, colWidths=[11*cm, 2.5*cm, 3.5*cm])
        diag_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,0), NAVY),
            ("TEXTCOLOR",     (0,0), (-1,0), WHITE),
            ("FONTSIZE",      (0,0), (-1,-1), 8),
            ("ROWBACKGROUNDS",(0,1), (-1,-1), [WHITE, LGREY]),
            ("GRID",          (0,0), (-1,-1), 0.4, MGREY),
            ("PADDING",       (0,0), (-1,-1), 6),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ]))
        story.append(Paragraph("Fault Distribution", sty("Normal", fontSize=10, textColor=NAVY, fontName="Helvetica-Bold", spaceAfter=4)))
        story.append(diag_tbl)

    # ══ SECTION 2: LOOP OVERVIEW ══════════════════════════════════════════════
    story.append(PageBreak())
    story.append(Paragraph("2. Loop Overview", H2))
    story.append(Paragraph(
        "All loops sorted from most critical to healthiest. "
        "Detailed analysis for each loop is in Section 3.", BODY))
    story.append(Spacer(1, 0.2*cm))

    # Header row
    ov_hdr = [
        Paragraph("<b>Loop</b>",      sty("Normal", fontSize=8, textColor=WHITE, spaceAfter=0)),
        Paragraph("<b>Severity</b>",  sty("Normal", fontSize=8, textColor=WHITE, spaceAfter=0)),
        Paragraph("<b>Health</b>",    sty("Normal", fontSize=8, textColor=WHITE, spaceAfter=0)),
        Paragraph("<b>Diagnosis</b>", sty("Normal", fontSize=8, textColor=WHITE, spaceAfter=0)),
        Paragraph("<b>Key Metric</b>",sty("Normal", fontSize=8, textColor=WHITE, spaceAfter=0)),
    ]
    ov_data = [ov_hdr]
    ov_ts   = [
        ("BACKGROUND",    (0,0), (-1,0), NAVY),
        ("FONTSIZE",      (0,0), (-1,-1), 8),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [WHITE, LGREY]),
        ("GRID",          (0,0), (-1,-1), 0.4, MGREY),
        ("PADDING",       (0,0), (-1,-1), 6),
        ("VALIGN",        (0,0), (-1,-1), "TOP"),
    ]

    for i, r in enumerate(sorted_loops, start=1):
        sev      = str(getattr(r, "Severity", "OK"))
        health_v = fmt(r._4, 0)
        diag_v   = str(r.Diagnosis) if str(r.Diagnosis) != "nan" else "—"
        harris_v = fmt(getattr(r, "Harris_Index", r._12 if hasattr(r,"_12") else "—"), 2)
        iae_v    = fmt(r._7 if hasattr(r,"_7") else "—", 0)
        key_met  = f"Harris: {harris_v}  |  IAE/hr: {iae_v}"
        sc = sev_color(sev)
        ov_data.append([
            Paragraph(str(r.Loop), MONO_CELL),
            Paragraph(f"<b>{sev}</b>", sty("Normal", fontSize=8, textColor=sc, fontName="Helvetica-Bold", spaceAfter=0)),
            Paragraph(f"<b>{health_v}</b>", sty("Normal", fontSize=8, textColor=sc, fontName="Helvetica-Bold", spaceAfter=0)),
            Paragraph(diag_v, CELL),
            Paragraph(key_met, sty("Normal", fontSize=7, textColor=DGREY, spaceAfter=0)),
        ])
        ov_ts.append(("BACKGROUND", (1,i), (2,i), sev_bg(sev)))

    ov_tbl = Table(ov_data, colWidths=[4.2*cm, 1.6*cm, 1.4*cm, 5.6*cm, 4.2*cm])
    ov_tbl.setStyle(TableStyle(ov_ts))
    story.append(ov_tbl)

    # ══ SECTION 3: PER-LOOP DETAILED ANALYSIS ════════════════════════════════
    story.append(PageBreak())
    story.append(Paragraph("3. Detailed Loop Analysis", H2))
    story.append(Paragraph(
        "Each loop is presented with its full diagnostic breakdown, key performance "
        "metrics, root-cause explanation, and recommended corrective action.", BODY))

    for r in sorted_loops:
        loop      = str(r.Loop)
        diag      = str(r.Diagnosis) if str(r.Diagnosis) != "nan" else "No fault detected"
        sev       = str(getattr(r, "Severity", "OK"))
        health_v  = fmt(r._4, 0)
        conf      = fmt(getattr(r, "Confidence", r._5 if hasattr(r,"_5") else "—"), 0)
        sf        = fmt(r._6  if hasattr(r,"_6")  else "—", 1)
        iae       = fmt(r._7  if hasattr(r,"_7")  else "—", 1)
        harris    = fmt(r._12 if hasattr(r,"_12") else "—", 3)
        op_act    = fmt(r._9  if hasattr(r,"_9")  else "—", 2)
        action    = str(r._22 if hasattr(r,"_22") else "—")
        rationale = str(r.Rationale) if hasattr(r, "Rationale") and str(r.Rationale) != "nan" else ""
        detail    = str(r._24 if hasattr(r,"_24") else "")
        if detail in ("nan", "None"): detail = ""

        sc  = sev_color(sev)
        sb  = sev_bg(sev)

        story.append(Spacer(1, 0.3*cm))

        # ── Loop header banner — always navy, severity as badge ───────────────
        sev_badge_color = "#e57373" if str(sev).upper() in ("FAIL","CRITICAL") else \
                          "#ffcc80" if str(sev).upper() in ("WARN","WARNING") else "#81c995"
        hdr_data = [[
            Paragraph(f"<font color='#ffffff'><b>{loop}</b></font>",
                      sty("Normal", fontSize=11, fontName="Helvetica-Bold", spaceAfter=0)),
            Paragraph(f"<font color='#a8c8f0'>{diag}</font>",
                      sty("Normal", fontSize=9, spaceAfter=0)),
            Paragraph(f"<font color='{sev_badge_color}'><b>{sev}</b></font>",
                      sty("Normal", fontSize=9, fontName="Helvetica-Bold", spaceAfter=0, alignment=2)),
        ]]
        hdr_tbl = Table(hdr_data, colWidths=[4.5*cm, 9.5*cm, 3*cm])
        hdr_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), NAVY),
            ("PADDING",    (0,0), (-1,-1), 9),
            ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
        ]))
        story.append(hdr_tbl)

        # ── Severity + health strip — plain light grey ───────────────────────
        sev_strip_data = [[
            Paragraph(f"Severity: <b><font color='#{to_hex(sc)}'>{sev}</font></b>",
                      sty("Normal", fontSize=8, textColor=NAVY, spaceAfter=0)),
            Paragraph(f"Health Score: <b>{health_v}/100</b>",
                      sty("Normal", fontSize=8, textColor=NAVY, spaceAfter=0)),
            Paragraph(f"Confidence: <b>{conf}%</b>",
                      sty("Normal", fontSize=8, textColor=NAVY, spaceAfter=0)),
        ]]
        sev_strip = Table(sev_strip_data, colWidths=[4*cm, 5*cm, 5*cm])
        sev_strip.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,-1), LGREY),
            ("PADDING",    (0,0), (-1,-1), 6),
            ("VALIGN",     (0,0), (-1,-1), "MIDDLE"),
            ("GRID",       (0,0), (-1,-1), 0.4, MGREY),
        ]))
        story.append(sev_strip)

        # ── Metrics grid ──────────────────────────────────────────────────────
        def metric_cell(label, val, unit=""):
            return [
                Paragraph(label, sty("Normal", fontSize=7, textColor=DGREY, spaceAfter=0)),
                Paragraph(f"<b>{val}{unit}</b>", sty("Normal", fontSize=10, textColor=NAVY, fontName="Helvetica-Bold", spaceAfter=0)),
            ]

        m_data = [
            [Paragraph("Service\nFactor", sty("Normal", fontSize=7, textColor=DGREY, spaceAfter=0)),
             Paragraph("Harris\nIndex",   sty("Normal", fontSize=7, textColor=DGREY, spaceAfter=0)),
             Paragraph("IAE / hr",        sty("Normal", fontSize=7, textColor=DGREY, spaceAfter=0)),
             Paragraph("OP Activity",     sty("Normal", fontSize=7, textColor=DGREY, spaceAfter=0))],
            [Paragraph(f"<b>{sf}%</b>",   sty("Normal", fontSize=12, textColor=NAVY, fontName="Helvetica-Bold", spaceAfter=0)),
             Paragraph(f"<b>{harris}</b>",sty("Normal", fontSize=12, textColor=NAVY, fontName="Helvetica-Bold", spaceAfter=0)),
             Paragraph(f"<b>{iae}</b>",   sty("Normal", fontSize=12, textColor=NAVY, fontName="Helvetica-Bold", spaceAfter=0)),
             Paragraph(f"<b>{op_act}</b>",sty("Normal", fontSize=12, textColor=NAVY, fontName="Helvetica-Bold", spaceAfter=0))],
        ]
        m_tbl = Table(m_data, colWidths=[4.25*cm]*4)
        m_tbl.setStyle(TableStyle([
            ("BACKGROUND",    (0,0), (-1,-1), LGREY),
            ("GRID",          (0,0), (-1,-1), 0.4, MGREY),
            ("PADDING",       (0,0), (-1,-1), 7),
            ("ALIGN",         (0,0), (-1,-1), "CENTER"),
            ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ]))
        story.append(m_tbl)

        # ── Summary ───────────────────────────────────────────────────────────
        if rationale:
            summary_data = [[
                Paragraph("📋 SUMMARY", sty("Normal", fontSize=7, textColor=DGREY, fontName="Helvetica-Bold", spaceAfter=2)),
            ],[
                Paragraph(rationale, BODY),
            ]]
            summary_tbl = Table(summary_data, colWidths=[17*cm])
            summary_tbl.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#f8fafc")),
                ("PADDING",    (0,0), (-1,-1), 8),
                ("LINEABOVE",  (0,0), (-1,0),  0.5, BLUE),
            ]))
            story.append(summary_tbl)

        # ── Detail explanation ────────────────────────────────────────────────
        if detail:
            detail_clean = (detail
                .replace("\\n\\n", "\n\n")
                .replace("\\n", " ")
                .replace("\n\n", "<br/><br/>")
                .replace("\n", " "))
            # Split into labelled sections (FINDING / ANALYSIS / CONCLUSION / etc.)
            sections = re.split(r'(FINDING:|ANALYSIS:|CONCLUSION:|IMPACT:|Oscillation Analysis\]|Performance Metrics\]|Controller Output\]|Process Variable\])', detail_clean)
            detail_rows = []
            i_s = 0
            while i_s < len(sections):
                chunk = sections[i_s].strip()
                if not chunk:
                    i_s += 1; continue
                if chunk.endswith(":") or chunk.endswith("]"):
                    label = chunk.lstrip("[")
                    body_chunk = sections[i_s+1].strip() if i_s+1 < len(sections) else ""
                    detail_rows.append([
                        Paragraph(f"<b>{label}</b>", sty("Normal", fontSize=7, textColor=BLUE, fontName="Helvetica-Bold", spaceAfter=0)),
                        Paragraph(body_chunk, BODY),
                    ])
                    i_s += 2
                else:
                    detail_rows.append(["", Paragraph(chunk, BODY)])
                    i_s += 1

            if detail_rows:
                det_tbl = Table(detail_rows, colWidths=[3*cm, 14*cm])
                det_tbl.setStyle(TableStyle([
                    ("BACKGROUND",   (0,0), (-1,-1), WHITE),
                    ("BACKGROUND",   (0,0), (0,-1), LGREY),
                    ("PADDING",      (0,0), (-1,-1), 6),
                    ("GRID",         (0,0), (-1,-1), 0.3, MGREY),
                    ("VALIGN",       (0,0), (-1,-1), "TOP"),
                    ("FONTSIZE",     (0,0), (-1,-1), 8),
                ]))
                story.append(det_tbl)

        # ── Recommended action ────────────────────────────────────────────────
        if action and action not in ("nan", "—", "None"):
            act_data = [[
                Paragraph("🔧 RECOMMENDED ACTION",
                          sty("Normal", fontSize=7, textColor=DGREY, fontName="Helvetica-Bold", spaceAfter=3)),
            ],[
                Paragraph(action, sty("Normal", fontSize=9, textColor=NAVY, leading=14, spaceAfter=0)),
            ]]
            act_tbl = Table(act_data, colWidths=[17*cm])
            act_tbl.setStyle(TableStyle([
                ("BACKGROUND",  (0,0), (-1,-1), colors.HexColor("#f8fafc")),
                ("PADDING",     (0,0), (-1,-1), 10),
                ("LINEBEFORE",  (0,0), (0,-1),  3, sc),
                ("LINEABOVE",   (0,0), (-1,0),   0.4, MGREY),
                ("LINEBELOW",   (0,-1),(-1,-1),  0.4, MGREY),
            ]))
            story.append(act_tbl)

        # ── Loop trend plot ───────────────────────────────────────────────────
        from reportlab.platypus import Image as RLImage
        plot_path = rd_path / "plots" / f"{loop}.png"
        if plot_path.exists():
            story.append(Spacer(1, 0.3*cm))
            trend_hdr = Table([[Paragraph(
                "📈 LOOP TREND (PV / SP / OP)",
                sty("Normal", fontSize=7, textColor=DGREY, fontName="Helvetica-Bold", spaceAfter=0)
            )]], colWidths=[17*cm])
            trend_hdr.setStyle(TableStyle([
                ("BACKGROUND", (0,0), (-1,-1), LGREY),
                ("PADDING",    (0,0), (-1,-1), 6),
                ("LINEABOVE",  (0,0), (-1,0),  0.4, MGREY),
            ]))
            story.append(trend_hdr)
            story.append(Spacer(1, 0.15*cm))
            try:
                story.append(RLImage(str(plot_path), width=17*cm, height=8*cm))
            except Exception:
                story.append(Paragraph("<i>[Trend plot could not be embedded]</i>",
                    sty("Normal", fontSize=8, textColor=DGREY, spaceAfter=0)))
            story.append(Spacer(1, 0.2*cm))

    # ══ SECTION 4: OUTLIER HANDLING ═══════════════════════════════════════════
    story.append(PageBreak())
    story.append(Paragraph("4. Data Quality — Outlier Handling", H2))
    story.append(Paragraph(
        "This section summarises how the tool handled anomalous or physically impossible "
        "values in the input data before running diagnostics. The original file was not modified.", BODY))
    story.append(Spacer(1, 0.2*cm))

    if outlier_text:
        for line in outlier_text.splitlines():
            line = line.strip()
            if not line:
                story.append(Spacer(1, 0.15*cm))
            elif any(line.startswith(k) for k in ("REMOVED", "FLAGGED")):
                story.append(Paragraph(line, sty("Normal", fontSize=9, textColor=CRIT, fontName="Helvetica-Bold")))
            elif any(k in line for k in ("KEPT", "No outliers")):
                story.append(Paragraph(line, sty("Normal", fontSize=9, textColor=OK, fontName="Helvetica-Bold")))
            elif any(line.startswith(k) for k in ("PRINCIPLES", "WHAT WE")):
                story.append(Paragraph(line, sty("Normal", fontSize=9, textColor=NAVY, fontName="Helvetica-Bold")))
            else:
                story.append(Paragraph(line, BODY))
    else:
        ok_box = Table([[Paragraph("✔  No outliers detected — input data was clean.", sty("Normal", fontSize=9, textColor=OK, fontName="Helvetica-Bold", spaceAfter=0))]],
                       colWidths=[17*cm])
        ok_box.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,-1),OK_BG),("PADDING",(0,0),(-1,-1),10)]))
        story.append(ok_box)

    # ══ INTERPRETATION GUIDE ══════════════════════════════════════════════════
    story.append(PageBreak())
    story.append(Paragraph("5. Interpretation Guide", H2))
    story.append(Paragraph("Plain-language explanation of every metric used in this report.", BODY))
    story.append(Spacer(1, 0.2*cm))

    notes = [
        ("Plant Health Index", "0–100 plant-wide score. ≥75 = Good. 50–74 = Needs attention. <50 = Critical."),
        ("Health Score (loop)", "0–100 per-loop score. 100 = no faults. Drops based on severity and confidence of detected faults."),
        ("Harris Index", "Control performance vs minimum achievable variance. <0.3 = poor. 0.3–0.7 = acceptable. >0.7 = good."),
        ("IAE / hr", "Integral Absolute Error per hour. How much PV deviates from SP over time. Lower is better."),
        ("Service Factor", "% of time the loop was in AUTO mode. <70% may indicate excessive manual intervention."),
        ("OP Activity", "Mean absolute change in controller output per sample. High values indicate oscillation or aggressive tuning."),
        ("Stiction", "Mechanical friction causing the valve to stick then jump. Diagnosed from the PV-OP relationship shape."),
        ("Saturation", "Valve at its physical limit (fully open/closed) and unable to respond further to controller demands."),
        ("Aggressive Tuning", "Controller gains too high, causing oscillation. PV and OP oscillate together with regularity >0.6."),
        ("Hägglund Regularity", "Measures how regular (periodic) the oscillation is. >0.6 = sustained oscillation."),
    ]
    note_hdr = [[
        Paragraph("<b>Metric</b>",      sty("Normal", fontSize=8, textColor=WHITE, spaceAfter=0)),
        Paragraph("<b>Explanation</b>", sty("Normal", fontSize=8, textColor=WHITE, spaceAfter=0)),
    ]]
    note_rows = note_hdr + [
        [Paragraph(f"<b>{t}</b>", sty("Normal", fontSize=8, textColor=NAVY, fontName="Helvetica-Bold", spaceAfter=0)),
         Paragraph(e, CELL)]
        for t, e in notes
    ]
    note_tbl = Table(note_rows, colWidths=[4.5*cm, 12.5*cm])
    note_tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0,0), (-1,0), NAVY),
        ("TEXTCOLOR",     (0,0), (-1,0), WHITE),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [WHITE, LGREY]),
        ("GRID",          (0,0), (-1,-1), 0.4, MGREY),
        ("PADDING",       (0,0), (-1,-1), 7),
        ("VALIGN",        (0,0), (-1,-1), "TOP"),
        ("FONTSIZE",      (0,0), (-1,-1), 8),
    ]))
    story.append(note_tbl)

    # ── Build ─────────────────────────────────────────────────────────────────
    doc.build(story)
    buf.seek(0)

    # ── Package PDF + Excel into a single zip ────────────────────────────────
    import zipfile
    ts      = datetime.now().strftime('%Y%m%d_%H%M')
    pdf_name = f"Valve_Diagnostic_Report_{ts}.pdf"
    zip_buf  = io.BytesIO()

    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        # Add the PDF we just built
        zf.writestr(pdf_name, buf.getvalue())
        # Add the styled PDF from disk (generated by engine.py)
        styled_pdf_path = rd_path / "Executive_summary_styled.pdf"
        if styled_pdf_path.exists():
            zf.write(styled_pdf_path, f"Valve_Diagnostic_Report_styled_{ts}.pdf")
        # Add the Excel from results folder
        if xl_path.exists():
            zf.write(xl_path, f"Loop_diagnostics_v2_{ts}.xlsx")

    zip_buf.seek(0)
    zip_name = f"Valve_Diagnostics_{ts}.zip"
    return StreamingResponse(
        zip_buf,
        media_type="application/zip",
        headers={"Content-Disposition": f"attachment; filename={zip_name}"}
    )




# ── Tuning Tool ───────────────────────────────────────────────────────────────

# In-memory store for tuning recommendations  { email: { loop_tag: { ...rec } } }
tuning_recs: dict[str, dict] = {}

@app.get("/tune/{loop_name:path}", response_class=HTMLResponse)
async def tune_page(loop_name: str, request: Request, current_user: dict = Depends(get_current_user)):
    """Serve the full-page tuning tool for a specific loop."""
    response = templates.TemplateResponse(request, "tune.html", {
        "loop_name": loop_name,
        "user": current_user,
    })
    response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
    return response


@app.post("/api/agent/tune")
async def agent_tune_analysis(request: Request, current_user: dict = Depends(get_current_user)):
    """AI agent that analyses why tuning was or wasn't possible for a loop."""
    body = await request.json()
    ctx = body.get("ctx", {})

    prompt = f"""You are an expert process control engineer analysing a control loop from a valve diagnostic tool.

Loop: {ctx.get('loop')}
Loop type: {ctx.get('loop_type')}
Diagnosis: {ctx.get('diagnosis_primary')} (Severity: {ctx.get('diagnosis_severity')})
Data samples: {ctx.get('n_samples')} (1-minute intervals = {round(ctx.get('n_samples',0)/60,1)} hours)

PROCESS MODEL IDENTIFICATION:
- Model selected: {ctx.get('model_type')}
- Process gain K = {ctx.get('K')} (method: {ctx.get('K_method')}, confidence: {ctx.get('conf_K_pct')}%, R2 = {ctx.get('r2_K')})
- Time constant tau = {ctx.get('tau')} min (method: {ctx.get('tau_method')}, confidence: {ctx.get('tau_conf_pct')}%)
- tau2 = {ctx.get('tau2') or 'N/A'}
- Dead time theta = {ctx.get('theta')} min
- Oscillation period Tu = {ctx.get('Tu')} min (confidence: {ctx.get('conf_Tu_pct')}%)
- OP range used: {ctx.get('op_range')}%
- No excitation flag: {ctx.get('no_excitation')}
- Model reason: {ctx.get('model_reason')} {ctx.get('model_reason_detail')}

TUNING RESULT:
- Tuning possible: {ctx.get('tuning_possible')}
- Lambda (IMC filter): {ctx.get('lambda')} min
- Current Kp (estimated): {ctx.get('Kp_old')} | Recommended Kp: {ctx.get('Kp_new')}
- Current Ti (estimated): {ctx.get('Ti_old')} min | Recommended Ti: {ctx.get('Ti_new')} min
- Predicted IAE improvement: {ctx.get('iae_improvement_pct')}%
- Predicted valve wear reduction: {ctx.get('op_wear_reduction_pct')}%

Write a clear, professional tuning analysis report for a control engineer. Structure it as:

1. TUNING STATUS - Was tuning possible? One sentence verdict.
2. WHY / WHY NOT - Explain the key reasons in plain English (2-4 sentences). Reference the actual numbers above.
3. MODEL CONFIDENCE - How reliable are K, tau, theta estimates? What does this mean for the recommendation?
4. RECOMMENDATION - IMPORTANT: If tuning_possible is False or Kp_new is null, do NOT suggest applying any Kp or Ti values. Instead give a concrete action plan for what the engineer must do first (e.g. fix the valve, perform an open-loop step test, put loop in manual). Only suggest Kp/Ti values if tuning_possible is True and Kp_new is not null.
5. WATCH OUT FOR - Any risks or things to verify.

Be concise, technical but readable. Use the section names as plain text labels. Do not invent numbers not provided above."""

    try:
        resp = _groq_chat(
            model=GROQ_MODEL,
            messages=[{"role": "user", "content": prompt}],
            max_tokens=800,
            temperature=0.3,
        )
        reply = resp.choices[0].message.content
        return JSONResponse(content={"reply": reply})
    except RuntimeError as e:
        raise HTTPException(status_code=500, detail=str(e))
    except Exception as e:
        msg = str(e)
        if "rate limit" in msg.lower() or "429" in msg:
            detail = "Agent is rate-limited right now. Please wait a minute and try again."
        else:
            detail = f"Agent unavailable: {msg[:200]}"
        raise HTTPException(status_code=503, detail=detail)


@app.get("/api/tune/timeseries/{loop_name:path}")
async def get_tune_timeseries(loop_name: str, current_user: dict = Depends(get_current_user)):
    """Return PV, OP, SP arrays for the tuning tool, sampled to max 1440 points."""
    rd = _get_user_results_dir(current_user)
    try:
        data = read_loop_timeseries(rd, loop_name, base_dir=str(BASE_DIR))
    except (FileNotFoundError, ValueError) as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    pv  = [v for v in data["pv"]  if v is not None]
    op  = [v for v in data["op"]  if v is not None]
    sp  = [v for v in data["sp"]  if v is not None]
    lbl = data["timestamps"]

    # Downsample to max 1440 points so the chart stays snappy
    N = len(pv)
    if N > 1440:
        step = N // 1440
        pv  = pv[::step][:1440]
        op  = op[::step][:1440]
        sp  = sp[::step][:1440]
        lbl = lbl[::step][:1440]

    # Detect loop type from tag name — handles codes embedded anywhere
    # e.g. YN.ETH1.16FC336 → FC → flow
    #      UN.UO.71PI1021C → PI → pressure
    import re as _re
    tag_up = loop_name.upper()

    # Method 1: digits+letters+digits pattern (most common in plant tags)
    _match = _re.search(r'\d+([A-Z]{1,3})\d+', tag_up)
    _code = _match.group(1) if _match else ''

    # Method 2: fallback — known codes as word/boundary match
    if not _code:
        for _c in ['FIC','FRC','FFC','PIC','PRC','TIC','TRC','LIC','LRC',
                   'FC','FI','FF','FT','PC','PI','PT','TC','TI','TT','LC','LI','LT']:
            if _re.search(r'(^|[^A-Z])' + _c + r'([^A-Z]|$)', tag_up):
                _code = _c
                break

    if _code in ('FC','FI','FF','FT','FY','FR','FQ','FIC','FRC','FFC'):
        loop_type = 'flow'
    elif _code in ('PC','PI','PT','PY','PR','PIC','PRC'):
        loop_type = 'pressure'
    elif _code in ('TC','TI','TT','TY','TR','TIC','TRC'):
        loop_type = 'temperature'
    elif _code in ('LC','LI','LT','LY','LR','LIC','LRC'):
        loop_type = 'level'
    else:
        loop_type = 'unknown'

    # ── Pull diagnosis from the Summary sheet so the tuning page knows
    # Look up diagnosis from this user's own results folder only
    diag_primary  = "unknown"
    diag_severity = "unknown"
    try:
        import pandas as _pd
        from pathlib import Path as _Path
        _target = loop_name.strip()
        _rd = _results_dir_for(current_user["sub"])
        if _rd:
            _xl = _Path(_rd) / "Loop_diagnostics_v2.xlsx"
            if _xl.exists():
                _sum = _pd.read_excel(str(_xl), sheet_name="Summary")
                _sum["Loop"] = _sum["Loop"].astype(str).str.strip()
                _row = _sum[_sum["Loop"] == _target]
                if not _row.empty:
                    diag_primary  = str(_row.iloc[0].get("Diagnosis",  "unknown") or "unknown").strip()
                    diag_severity = str(_row.iloc[0].get("Severity",   "unknown") or "unknown").strip()
    except Exception:
        pass  # non-fatal — tuning page degrades gracefully if lookup fails

    return JSONResponse(content={
        "loop": loop_name,
        "loop_type": loop_type,
        "n_samples": len(pv),
        "pv":  pv,
        "op":  op,
        "sp":  sp,
        "labels": lbl,
        "diag_primary":  diag_primary,
        "diag_severity": diag_severity,
    })


@app.post("/api/tune/save/{loop_name:path}")
async def save_tuning_rec(loop_name: str, request: Request, current_user: dict = Depends(get_current_user)):
    """Save an accepted tuning recommendation for a loop."""
    body = await request.json()
    from datetime import datetime
    email = current_user["sub"]
    if email not in tuning_recs:
        tuning_recs[email] = {}
    tuning_recs[email][loop_name] = {
        "loop":       loop_name,
        "loop_type":  body.get("loop_type", "unknown"),
        "K":          body.get("K"),
        "tau":        body.get("tau"),
        "theta":      body.get("theta"),
        "Tu":         body.get("Tu"),
        "lambda":     body.get("lambda"),
        "Kp_old":     body.get("Kp_old"),
        "Ti_old":     body.get("Ti_old"),
        "Kp_new":     body.get("Kp_new"),
        "Ti_new":     body.get("Ti_new"),
        "conf_K":     body.get("conf_K"),
        "r2":         body.get("r2"),
        "iae_pct":    body.get("iae_pct"),
        "status":     body.get("status", "accepted"),  # accepted | rejected | pending
        "notes":      body.get("notes", ""),
        "saved_at":   datetime.now().isoformat(timespec="seconds"),
    }
    return JSONResponse(content={"status": "ok", "loop": loop_name})


@app.get("/api/tune/recommendations")
async def get_tuning_recs(current_user: dict = Depends(get_current_user)):
    """Return all saved tuning recommendations for the current user."""
    email = current_user["sub"]
    recs  = tuning_recs.get(email, {})
    return JSONResponse(content={"recommendations": list(recs.values())})


@app.get("/api/tune/recommendation/{loop_name:path}")
async def get_tuning_rec(loop_name: str, current_user: dict = Depends(get_current_user)):
    """Return the saved tuning recommendation for a specific loop, if any."""
    email = current_user["sub"]
    rec   = tuning_recs.get(email, {}).get(loop_name)
    if not rec:
        return JSONResponse(content={"recommendation": None})
    return JSONResponse(content={"recommendation": rec})


@app.exception_handler(HTTPException)
async def http_exception_handler(request: Request, exc: HTTPException):
    if exc.status_code == 401:
        return RedirectResponse(url="/login", status_code=302)
    return JSONResponse(status_code=exc.status_code, content={"detail": exc.detail})


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8001, reload=True)