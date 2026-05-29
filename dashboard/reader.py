"""
reader.py - reads the latest valve diagnostics results folder
"""
import os
import glob
from datetime import datetime
from typing import Optional
import openpyxl


def read_unit_mapping(results_dir: str, base_dir: str) -> dict:
    """
    Read UNIT_MAPPING sheet from the original input Excel file.
    Derives filename from results folder name: results_<stem> -> <stem>.xlsx
    Returns {unit_map: {tag: unit}, units: [sorted unique units]}
    """
    folder_name = os.path.basename(results_dir)
    stem = folder_name.removeprefix("results_").removesuffix("_manual")
    input_path = os.path.join(base_dir, f"{stem}.xlsx")
    if not os.path.exists(input_path):
        return {"unit_map": {}, "units": [], "uom_map": {}}
    try:
        wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
        if "UNIT_MAPPING" not in wb.sheetnames:
            wb.close()
            return {"unit_map": {}, "units": [], "uom_map": {}}
        ws = wb["UNIT_MAPPING"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        unit_map = {}
        uom_map = {}
        for row in rows[1:]:
            if row[0] and row[1]:
                unit_map[str(row[0]).strip()] = str(row[1]).strip()
            if row[0] and len(row) > 2 and row[2]:
                uom_map[str(row[0]).strip()] = str(row[2]).strip()
        units = sorted(set(unit_map.values()))
        return {"unit_map": unit_map, "units": units, "uom_map": uom_map}
    except Exception:
        return {"unit_map": {}, "units": [], "uom_map": {}}


def read_unit_mapping_as_list(results_dir: str, base_dir: str) -> list:
    """
    Read UNIT_MAPPING sheet and return as list of {loop, unit} dicts
    for use in config JSON (unit_mapping key).
    """
    data = read_unit_mapping(results_dir, base_dir)
    return [
        {"loop": loop, "unit": unit}
        for loop, unit in data["unit_map"].items()
    ]


def find_latest_results_dir(base_dir: str) -> Optional[str]:
    pattern = os.path.join(base_dir, "results_*")
    folders = glob.glob(pattern)
    if not folders:
        return None
    folders.sort(key=lambda f: os.path.getmtime(f), reverse=True)
    return folders[0]


def _read_detection_exclusions(base_dir: str, results_dir: str) -> dict:
    """
    Read DETECTION_EXCLUSIONS sheet from the original input Excel.
    Returns {problem_id: set(loop_names)} e.g. {'stiction': {'LOOP_A', 'LOOP_B'}}
    """
    folder_name = os.path.basename(results_dir)
    stem = folder_name.removeprefix("results_").removesuffix("_manual")
    input_path = os.path.join(base_dir, f"{stem}.xlsx")
    if not os.path.exists(input_path):
        return {}
    try:
        wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
        if "DETECTION_EXCLUSIONS" not in wb.sheetnames:
            wb.close()
            return {}
        ws = wb["DETECTION_EXCLUSIONS"]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        excl = {}
        for row in rows[1:]:
            if row[0] and row[1]:
                pid = str(row[0]).strip()
                loop = str(row[1]).strip()
                if pid not in excl:
                    excl[pid] = set()
                excl[pid].add(loop)
        return excl
    except Exception:
        return {}


def _apply_detection_exclusions(loops: list, excl: dict) -> list:
    """
    For each loop, if its diagnosis matches an excluded problem for that loop,
    suppress that diagnosis (set to 'Normal' or next non-excluded diagnosis).
    excl = {problem_id: set(loop_names)}
    Problem IDs use underscores: 'stiction', 'aggressive_tuning', etc.
    """
    if not excl:
        return loops

    # Map diagnosis string fragments to problem IDs
    # Fragments are matched as substrings (case-insensitive) against the diagnosis field
    DIAG_TO_PROBLEM = {
        'stiction':                  'stiction',
        'aggressive tuning':         'aggressive_tuning',
        'aggressive_tuning':         'aggressive_tuning',
        'sluggish tuning':           'sluggish_tuning',
        'sluggish_tuning':           'sluggish_tuning',
        'external oscillation':      'external_oscillation',
        'external_oscillation':      'external_oscillation',
        'cross-loop propagation':    'cross_loop_propagation',
        'cross_loop_propagation':    'cross_loop_propagation',
        'sensor noise':              'sensor_noise',
        'sensor_noise':              'sensor_noise',
        'sensor issue':              'sensor_noise',
        'manual mode':               'manual_mode',
        'manual_mode':               'manual_mode',
        'loop in man':               'manual_mode',
        'valve wear':                'valve_wear',
        'valve_wear':                'valve_wear',
        'saturation':                'saturation',
        'unresponsive controller':   'unresponsive_controller',
        'unresponsive':              'unresponsive_controller',
        'oscillation':               'aggressive_tuning',
    }

    result = []
    for loop in loops:
        loop = dict(loop)  # copy so we don't mutate original
        loop_name = loop.get('loop', '')
        diag = loop.get('diagnosis', '')
        diag_lower = diag.lower().strip()

        # Find which problem_id this diagnosis maps to
        matched_pid = None
        for fragment, pid in DIAG_TO_PROBLEM.items():
            if fragment in diag_lower:
                matched_pid = pid
                break

        # If this loop is excluded for its current diagnosis, clear it
        if matched_pid and loop_name in excl.get(matched_pid, set()):
            loop['diagnosis'] = 'Normal'
            loop['severity'] = 'OK'
            loop['health'] = 100.0
            loop['recommended_action'] = ''
            loop['rationale'] = ''

        result.append(loop)
    return result


def read_dashboard_data(results_dir: str, base_dir: str = None) -> dict:
    xlsx_path = os.path.join(results_dir, "Loop_diagnostics_v2.xlsx")
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"No Loop_diagnostics_v2.xlsx in {results_dir}")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    plant = _read_plant_dashboard(wb)
    loops = _read_summary(wb)
    stiction = _read_stiction(wb)
    propagation = _read_propagation(wb)
    data_quality = _read_data_quality(wb)
    maintenance = _read_maintenance(wb)

    stiction_map = {r["loop"]: r for r in stiction}
    dq_map = {r["loop"]: r for r in data_quality}
    for loop in loops:
        name = loop["loop"]
        if name in stiction_map:
            loop["stiction"] = stiction_map[name]
        if name in dq_map:
            loop["data_quality"] = dq_map[name]

    # Apply detection exclusions if base_dir is available
    if base_dir:
        excl = _read_detection_exclusions(base_dir, results_dir)
        loops = _apply_detection_exclusions(loops, excl)

    run_summary = _read_run_summary(results_dir)
    wb.close()
    return {
        "run_folder": os.path.basename(results_dir),
        "run_summary": run_summary,
        "plant": plant,
        "loops": loops,
        "propagation": propagation,
        "maintenance": maintenance,
    }


def read_loop_timeseries(results_dir: str, loop_name: str) -> dict:
    """Read PV, OP, SP, MODE time series for a specific loop from data_v3_processed.xlsx."""
    xlsx_path = os.path.join(results_dir, "data_v3_processed.xlsx")
    if not os.path.exists(xlsx_path):
        raise FileNotFoundError(f"No data_v3_processed.xlsx in {results_dir}")

    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        raise ValueError("No data found in Sheet1")

    headers = [str(c).strip() if c is not None else "" for c in rows[0]]

    # Find columns for this loop — try exact match and all separator variants
    SEPARATORS = ["_", "-", ".", " "]
    SUFFIXES = ["PV", "SP", "OP", "MODE"]

    def find_col(suffix):
        suffix_up = suffix.upper()
        # 1. Exact: loop_name + sep + suffix (any case)
        for sep in SEPARATORS:
            for s in [suffix_up, suffix.lower(), suffix]:
                candidate = f"{loop_name}{sep}{s}"
                if candidate in headers:
                    return headers.index(candidate)
        # 2. No separator: loop_name + suffix directly
        for s in [suffix_up, suffix.lower(), suffix]:
            candidate = f"{loop_name}{s}"
            if candidate in headers:
                return headers.index(candidate)
        # 3. Case-insensitive scan: header ends with any sep+suffix or just suffix
        for i, h in enumerate(headers):
            hu = h.upper()
            # with separator
            for sep in SEPARATORS:
                if hu.endswith(sep + suffix_up):
                    loop_part = h[:-(len(suffix) + 1)]
                    if loop_part.replace("-","_").replace(".","_") == loop_name.replace("-","_").replace(".","_"):
                        return i
            # without separator
            if hu.endswith(suffix_up) and not any(hu.endswith(sep + suffix_up) for sep in SEPARATORS):
                loop_part = h[:-len(suffix)]
                if loop_part.replace("-","_").replace(".","_") == loop_name.replace("-","_").replace(".","_"):
                    return i
        return None

    ts_idx = 0  # TIMESTAMP always first
    pv_idx = find_col("PV")
    op_idx = find_col("OP")
    sp_idx = find_col("SP")
    mode_idx = find_col("MODE")

    if pv_idx is None:
        # Try to find any column containing the loop name
        available = [h for h in headers if loop_name.replace("-","_") in h.replace("-","_")]
        raise ValueError(f"Loop '{loop_name}' not found. Available columns: {available[:10]}")

    timestamps = []
    pv_vals = []
    op_vals = []
    sp_vals = []
    mode_vals = []

    for row in rows[1:]:
        ts = row[ts_idx]
        if ts is None:
            continue
        # Format timestamp
        if hasattr(ts, 'strftime'):
            ts_str = ts.strftime("%Y-%m-%d %H:%M")
        else:
            ts_str = str(ts)

        timestamps.append(ts_str)
        pv_vals.append(_safe_float(row[pv_idx]))
        op_vals.append(_safe_float(row[op_idx]) if op_idx is not None else None)
        sp_vals.append(_safe_float(row[sp_idx]) if sp_idx is not None else None)
        mode_vals.append(str(row[mode_idx]).strip() if mode_idx is not None and row[mode_idx] is not None else "")

    return {
        "loop": loop_name,
        "timestamps": timestamps,
        "pv": pv_vals,
        "op": op_vals,
        "sp": sp_vals,
        "mode": mode_vals,
        "columns_found": {
            "pv": headers[pv_idx] if pv_idx is not None else None,
            "op": headers[op_idx] if op_idx is not None else None,
            "sp": headers[sp_idx] if sp_idx is not None else None,
            "mode": headers[mode_idx] if mode_idx is not None else None,
        }
    }


def read_all_loop_names(results_dir: str) -> list:
    """Return list of loop names found in data_v3_processed.xlsx."""
    xlsx_path = os.path.join(results_dir, "data_v3_processed.xlsx")
    if not os.path.exists(xlsx_path):
        return []
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb["Sheet1"]
    headers = [str(c).strip() if c is not None else "" for c in next(ws.iter_rows(values_only=True))]
    wb.close()
    loops = set()
    for h in headers:
        if h.upper() in ('TIMESTAMP', ''):
            continue
        matched = False
        for sep in ['_', '-', '.', ' ']:
            for suffix in ['PV', 'SP', 'OP', 'MODE']:
                if h.upper().endswith(sep + suffix):
                    loops.add(h[:-(len(suffix) + 1)])
                    matched = True
                    break
            if matched:
                break
        if not matched:
            for suffix in ['PV', 'SP', 'OP', 'MODE']:
                if h.upper().endswith(suffix) and not any(h.upper().endswith(sep + suffix) for sep in ['_', '-', '.', ' ']):
                    loops.add(h[:-len(suffix)])
                    break
    return sorted(loops)


def _sheet_rows(wb, sheet_name: str):
    if sheet_name not in wb.sheetnames:
        return [], []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    return header, rows[1:]


def _read_plant_dashboard(wb) -> dict:
    ws = wb["Plant_Dashboard"]
    rows = list(ws.iter_rows(values_only=True))
    kv = {}
    diagnosis_counts = {}
    loop_health = []
    section = "kv"
    for row in rows[1:]:
        if row[0] == "Diagnosis":
            section = "diagnosis"
            continue
        if row[0] == "Loop":
            section = "loops"
            continue
        if section == "kv" and row[0] and row[1] is not None:
            kv[str(row[0]).strip()] = row[1]
        elif section == "diagnosis" and row[0] and row[1] is not None:
            diagnosis_counts[str(row[0]).strip()] = int(row[1])
        elif section == "loops" and row[0]:
            loop_health.append({
                "loop": str(row[0]).strip(),
                "health": float(row[1]) if row[1] is not None else 0,
                "diagnosis": str(row[2]).strip() if row[2] else ""
            })
    return {
        "plant_health_index": float(kv.get("Plant Health Index (0-100)", 0)),
        "loops_total": int(kv.get("Loops total", 0)),
        "loops_analysed": int(kv.get("Loops analysed", 0)),
        "loops_skipped": int(kv.get("Loops skipped", 0)),
        "pct_good": float(kv.get("% Good (>=75)", 0)),
        "pct_poor": float(kv.get("% Poor (50\u201374)", 0)),
        "pct_critical": float(kv.get("% Critical (<50)", 0)),
        "sample_interval": str(kv.get("Sample interval", "")),
        "duration_hours": float(kv.get("Duration (hours)", 0)),
        "run_timestamp": str(kv.get("Run timestamp", "")),
        "diagnosis_counts": diagnosis_counts,
        "loop_health_ranking": loop_health,
    }


def _read_summary(wb) -> list:
    header, rows = _sheet_rows(wb, "Summary")
    results = []
    for row in rows:
        if not row[0]:
            continue
        d = dict(zip(header, row))
        results.append({
            "loop": str(d.get("Loop", "")).strip(),
            "diagnosis": str(d.get("Diagnosis", "")).strip(),
            "severity": str(d.get("Severity", "")).strip(),
            "health": _safe_float(d.get("Health (0-100)")),
            "confidence": _safe_float(d.get("Confidence")),
            "service_factor": _safe_float(d.get("Service Factor %")),
            "iae_per_hour": _safe_float(d.get("IAE/hr")),
            "pv_amplitude": _safe_float(d.get("PV amplitude")),
            "op_activity": _safe_float(d.get("OP activity")),
            "harris_index": _safe_float(d.get("Harris Index")),
            "hagglund_regularity": _safe_float(d.get("H\u00e4gglund regularity")),
            "dominant_period": str(d.get("Dominant period", "") or ""),
            "stiction_label": str(d.get("Stiction label", "") or "").strip(),
            "recommended_action": str(d.get("Recommended action", "") or "").strip(),
            "rationale": str(d.get("Rationale", "") or "").strip(),
            "data_quality_status": str(d.get("Data quality", "") or "").strip(),
            "issues": str(d.get("Issues", "") or "").strip(),
        })
    return results


def _read_stiction(wb) -> list:
    header, rows = _sheet_rows(wb, "Stiction_Analysis")
    results = []
    for row in rows:
        if not row[0]:
            continue
        d = dict(zip(header, row))
        results.append({
            "loop": str(d.get("Loop", "")).strip(),
            "heuristic": _safe_float(d.get("Heuristic")),
            "horch_cc": _safe_float(d.get("Horch CC")),
            "yamashita": _safe_float(d.get("Yamashita shape")),
            "bicoherence": _safe_float(d.get("Bicoherence")),
            "methods_agreeing": _safe_float(d.get("Methods agreeing (>50)")),
            "consensus": _safe_float(d.get("Consensus")),
            "label": str(d.get("Label", "") or "").strip(),
            "estimated_s": _safe_float(d.get("Estimated S (%)")),
            "estimated_j": _safe_float(d.get("Estimated J (%)")),
            "shape": str(d.get("Shape", "") or "").strip(),
        })
    return results


def _read_propagation(wb) -> list:
    header, rows = _sheet_rows(wb, "Propagation")
    results = []
    for row in rows:
        if not row[0]:
            continue
        d = dict(zip(header, row))
        results.append({
            "source": str(d.get("Source loop", "")).strip(),
            "target": str(d.get("Target loop", "")).strip(),
            "combined_score": _safe_float(d.get("Combined score")),
            "lag_time": str(d.get("Lag (time)", "") or ""),
            "cross_correlation": _safe_float(d.get("Cross-correlation")),
            "coherence_score": _safe_float(d.get("Coherence score")),
        })
    return [r for r in results if (r["combined_score"] or 0) >= 50]


def _read_data_quality(wb) -> list:
    header, rows = _sheet_rows(wb, "Per_Loop_Data_Quality")
    results = []
    for row in rows:
        if not row[0]:
            continue
        d = dict(zip(header, row))
        results.append({
            "loop": str(d.get("Loop", "")).strip(),
            "samples": _safe_int(d.get("Samples")),
            "pct_missing": _safe_float(d.get("% missing")),
            "quantised": str(d.get("Quantised?", "") or "").strip(),
            "frozen": str(d.get("Frozen?", "") or "").strip(),
            "outliers": _safe_int(d.get("Outliers")),
            "severity": str(d.get("Severity", "") or "").strip(),
            "issues": str(d.get("Issues", "") or "").strip(),
        })
    return results


def _read_maintenance(wb) -> list:
    header, rows = _sheet_rows(wb, "Maintenance_Actions")
    results = []
    for row in rows:
        if not row[0]:
            continue
        d = dict(zip(header, row))
        results.append({
            "loop": str(d.get("Loop", "")).strip(),
            "severity": str(d.get("Severity", "")).strip(),
            "diagnosis": str(d.get("Diagnosis", "")).strip(),
            "health": _safe_float(d.get("Health")),
            "recommended_action": str(d.get("Recommended action", "") or "").strip(),
            "confidence": _safe_float(d.get("Confidence")),
        })
    results.sort(key=lambda x: x["health"] or 100)
    return results


def _read_run_summary(results_dir: str) -> dict:
    summary_path = os.path.join(results_dir, "v3_run_summary.txt")
    if not os.path.exists(summary_path):
        return {}
    data = {}
    with open(summary_path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if ":" in line and not line.startswith("\u2550") and not line.startswith("\u2500"):
                key, _, val = line.partition(":")
                data[key.strip()] = val.strip()
    return data


def _safe_float(val) -> Optional[float]:
    try:
        return float(val) if val is not None and val != "" else None
    except (ValueError, TypeError):
        return None


def _safe_int(val) -> Optional[int]:
    try:
        return int(val) if val is not None and val != "" else None
    except (ValueError, TypeError):
        return None


def list_all_runs(base_dir: str) -> list:
    pattern = os.path.join(base_dir, "results_*")
    folders = glob.glob(pattern)
    folders.sort(key=lambda f: os.path.getmtime(f), reverse=True)
    return [
        {
            "name": os.path.basename(f),
            "modified": datetime.fromtimestamp(os.path.getmtime(f)).strftime("%Y-%m-%d %H:%M:%S"),
            "path": f,
        }
        for f in folders
    ]
