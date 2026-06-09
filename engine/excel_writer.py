"""
Excel writer — builds the Loop_diagnostics_v2.xlsx output workbook.
=====================================================================

Layer 4, module 2. Produces the multi-sheet Excel report:

  * Summary
  * Per-loop diagnoses
  * Diagnostic configuration (with documentation)
  * Propagation links
  * Plant KPIs
"""

import os
from datetime import datetime
from typing import List

import numpy as np
import pandas as pd

from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .utils import logger, safe_float, safe_pos, DEFAULTS
from .time_context import TimeContext
from .capabilities import Capabilities
from .plant_kpis import PlantKPIs


def _bold_header(ws, header_row=1):
    bold = Font(bold=True)
    fill = PatternFill("solid", fgColor="DDDDDD")
    for cell in ws[header_row]:
        cell.font = bold
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws):
    for col in ws.columns:
        try:
            max_len = max(len(str(c.value)) for c in col if c.value is not None)
        except ValueError:
            max_len = 12
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)


def write_excel_report(path: str, df: pd.DataFrame, tc: TimeContext,
                       capabilities: Capabilities, per_loop: dict,
                       links: list, kpi: PlantKPIs, config: dict,
                       dashboard_png: str = None,
                       selection=None,
                       heatmap_png: str = None):
    """Write the full Excel diagnostic workbook."""
    # Build summary table first
    rows = []
    for name, info in per_loop.items():
        m = info.get("metrics")
        d = info.get("diagnosis")
        s = info.get("stiction")
        dq = info.get("data_quality")
        sf = info.get("service_factor", 100.0)
        hi = info.get("harris", float("nan"))
        if d is None:
            rows.append({"Loop": name, "Status": "SKIPPED",
                         "Reason": info.get("skip_reason", "n/a")})
            continue
        rows.append({
            "Loop": name,
            "Diagnosis": d.primary,
            "Severity": d.severity,
            "Health (0-100)": d.health_score,
            "Confidence": round(d.confidence, 1),
            "Service Factor %": sf,
            "IAE/hr": m.iae_per_hour if m else None,
            "IAE/hr normalised %": round(m.iae_per_hour_norm, 1) if m else None,
            "PV amplitude": round(m.pv_amplitude, 2) if m else None,
            "OP activity": round(m.op_activity, 2) if m else None,
            "OP% at 0%": m.op_pct_at_zero if m else None,
            "OP% at full": m.op_pct_at_full if m else None,
            "Harris Index": round(hi, 3) if not np.isnan(hi) else None,
            "Hägglund regularity": info.get("osc_reg", None),
            "Dominant period": info.get("osc_period_str", None),
            "Stiction consensus": s.consensus_score if s else None,
            "Stiction label": s.consensus_label if s else None,
            "Stiction S (% OP)": s.estimated_S if s else None,
            "Stiction J (% OP)": s.estimated_J if s else None,
            "Yamashita shape": s.yamashita_shape if s else None,
            "Data quality": dq.severity if dq else None,
            "Issues": "; ".join(dq.issues) if dq and dq.issues else "",
            "Recommended action": d.recommended_action,
            "Rationale": d.rationale,
            "Detailed Explanation": d.detailed_explanation,
        })
    summary_df = pd.DataFrame(rows)

    # Plant dashboard table
    dash_df = pd.DataFrame({
        "Metric": ["Plant Health Index (0-100)", "Loops total", "Loops analysed",
                   "Loops skipped", "% Good (>=75)", "% Poor (50–74)",
                   "% Critical (<50)", "Sample interval", "Duration (hours)",
                   "Run timestamp"],
        "Value": [kpi.plant_health_index, kpi.n_loops_total,
                  kpi.n_loops_analysed, kpi.n_skipped,
                  kpi.pct_good, kpi.pct_poor, kpi.pct_critical,
                  tc.dt_str(), round(tc.duration_hours, 2),
                  datetime.now().isoformat(timespec="seconds")],
    })

    diag_dist_df = pd.DataFrame(
        list(kpi.diagnosis_counts.items()),
        columns=["Diagnosis", "Count"],
    ).sort_values("Count", ascending=False)

    top_worst_df = pd.DataFrame(
        [(n, h, d) for n, h, d in kpi.top_n_worst],
        columns=["Loop", "Health", "Primary diagnosis"],
    )

    # Capability / coverage sheet
    cov_rows = []
    cov_rows.append({"Diagnostic": "Sample-rate", "Status": "DETECTED",
                     "Detail": f"interval={tc.dt_str()}, duration={tc.duration_hours:.2f}h, "
                               f"irregular={tc.irregular}, gaps={tc.gap_count}"})
    for diag_name, can_run in [("Stiction", capabilities.can_stiction),
                               ("Oscillation", capabilities.can_oscillation),
                               ("Harris index", capabilities.can_harris),
                               ("Propagation", capabilities.can_propagation)]:
        if can_run:
            cov_rows.append({"Diagnostic": diag_name, "Status": "ENABLED", "Detail": ""})
        else:
            key = diag_name.lower().split()[0]
            reason = capabilities.skip_reasons.get(key, "Skipped")
            cov_rows.append({"Diagnostic": diag_name, "Status": "DISABLED", "Detail": reason})
    coverage_df = pd.DataFrame(cov_rows)

    # Stiction analysis detail
    stic_rows = []
    for name, info in per_loop.items():
        s = info.get("stiction")
        if s is None:
            continue
        stic_rows.append({
            "Loop": name,
            "Heuristic": s.heuristic_score,
            "Horch CC": s.horch_score,
            "Yamashita shape": s.yamashita_score,
            "Bicoherence": s.bicoherence_score,
            "Methods agreeing (>50)": s.methods_agreeing,
            "Consensus": s.consensus_score,
            "Label": s.consensus_label,
            "Estimated S (%)": s.estimated_S,
            "Estimated J (%)": s.estimated_J,
            "Shape": s.yamashita_shape,
        })
    stiction_df = pd.DataFrame(stic_rows)

    # Propagation
    prop_rows = []
    for l in links:
        prop_rows.append({
            "Source loop": l.source,
            "Source unit": l.source_unit,
            "Target loop": l.target,
            "Target unit": l.target_unit,
            "Same unit?": "Yes" if l.same_unit else "No (cross-unit)",
            "Cross-correlation": l.cc_correlation,
            "Lag (samples)": l.cc_lag_samples,
            "Lag (time)": l.cc_lag_str,
            "Granger p-value": l.granger_p,
            "Coherence score": l.coherence_score,
            "Raw score": l.raw_score,
            "Combined score": l.combined_score,
        })
    prop_df = pd.DataFrame(prop_rows)

    # Data quality detail
    dq_rows = []
    for name, info in per_loop.items():
        dq = info.get("data_quality")
        if dq is None:
            continue
        dq_rows.append({
            "Loop": name,
            "Samples": dq.n_samples,
            "Finite": dq.n_finite,
            "% missing": dq.pct_missing,
            "Unique PV values": dq.pv_unique_values,
            "Quantised?": dq.is_quantised,
            "Longest frozen run": dq.longest_frozen_run,
            "Frozen?": dq.is_frozen,
            "Compression fraction": dq.compression_fraction,
            "Compressed?": dq.is_compressed,
            "Outliers": dq.n_outliers,
            "Severity": dq.severity,
            "Issues": "; ".join(dq.issues) if dq.issues else "",
        })
    dq_df = pd.DataFrame(dq_rows)

    # Maintenance actions
    maint_rows = []
    for name, info in per_loop.items():
        d = info.get("diagnosis")
        if d is None:
            continue
        if d.severity in ("WARN", "FAIL"):
            maint_rows.append({
                "Loop": name,
                "Severity": d.severity,
                "Diagnosis": d.primary,
                "Health": d.health_score,
                "Recommended action": d.recommended_action,
                "Confidence": round(d.confidence, 1),
                "Detailed Explanation": d.detailed_explanation,
            })
    maint_df = pd.DataFrame(maint_rows).sort_values(
        ["Severity", "Health"], ascending=[False, True]
    ) if maint_rows else pd.DataFrame()

    # How-to-read sheet
    how_to = pd.DataFrame({
        "Term": [
            "Plant Health Index (PHI)",
            "Health score (per loop)",
            "Service Factor",
            "Harris Index",
            "Hägglund regularity",
            "IAE per hour",
            "OP activity",
            "Stiction consensus",
            "Stiction S (% OP)",
            "Stiction J (% OP)",
            "Yamashita shape",
            "Granger p-value",
            "Spectral coherence",
            "Sample-rate gating",
        ],
        "Plain-language meaning": [
            "Average of all loop health scores. >=75 is healthy plant; <50 is critical.",
            "0-100 score; 100 means no detected issues. Fault rules deduct.",
            "% of time the loop was in AUTO/CAS. <70% = operator overrides too often.",
            "0–1; ratio of theoretical minimum variance to actual variance. >0.5 is decent.",
            "0–1 measure of how regular the PV oscillations are. >0.6 = clear oscillation.",
            "Time-averaged tracking error per hour. Lower is better.",
            "Average |dOP| between consecutive samples. High value = valve hunting.",
            "0–100 weighted score from 4 stiction tests. >70 + 2 methods = Confirmed.",
            "Estimated stickband — OP must change by this much before valve moves.",
            "Estimated slip-jump — how far valve overshoots when it breaks free.",
            "Visual classification of PV-OP plot: straight=healthy, parallelogram=stiction.",
            "Statistical evidence that one PV drives another. <0.05 = significant.",
            "0–100; how strongly two loops oscillate at the same frequency.",
            "Some diagnostics need fast data; tool skips them on coarse data and reports why.",
        ],
    })

    # Config documentation
    cfg_doc_rows = [{"Parameter": k, "Value": v,
                     "Description": _CONFIG_DOC.get(k, "")} for k, v in config.items()]
    cfg_doc_df = pd.DataFrame(cfg_doc_rows)

    # Diagnostic Selection summary (which diagnoses ran, and why some didn't)
    sel_rows = []
    if selection is not None:
        # Parents
        parents = [
            ("Stiction detection", selection.stiction_detection),
            ("Aggressive tuning", selection.aggressive_tuning),
            ("Sluggish tuning", selection.sluggish_tuning),
            ("External oscillation", selection.external_oscillation),
            ("Cross-loop propagation", selection.cross_loop_propagation),
        ]
        for label, on in parents:
            sel_rows.append({"Diagnostic": label, "Type": "Parent",
                             "Status": "RAN" if on else "DISABLED"})
        # Methods
        methods = [
            ("Heuristic", selection.stic_heuristic),
            ("Horch cross-correlation", selection.stic_horch),
            ("Yamashita shape", selection.stic_yamashita),
            ("Bicoherence", selection.stic_bicoherence),
            ("Fall back to other indicators", selection.stic_fall_back),
            ("Harris Index", selection.harris_index),
            ("Hägglund oscillation", selection.hagglund_oscillation),
            ("Cross-correlation (propagation)", selection.prop_cross_correlation),
            ("Granger causality", selection.prop_granger),
            ("Spectral coherence", selection.prop_coherence),
        ]
        for label, on in methods:
            sel_rows.append({"Diagnostic": label, "Type": "Method",
                             "Status": "RAN" if on else "DISABLED"})
        # Auto-disable explanations
        for msg in selection.auto_disable_log:
            sel_rows.append({"Diagnostic": "(auto-disable)", "Type": "Note",
                             "Status": msg})
    sel_df = pd.DataFrame(sel_rows)

    # Determine which sheets are skippable based on user selection
    show_stiction_sheet = (selection is None) or selection.stiction_detection
    show_propagation_sheet = (selection is None) or selection.cross_loop_propagation

    # Write workbook
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        dash_df.to_excel(writer, sheet_name="Plant_Dashboard", index=False, startrow=0)
        diag_dist_df.to_excel(writer, sheet_name="Plant_Dashboard", index=False,
                              startrow=len(dash_df) + 3)
        top_worst_df.to_excel(writer, sheet_name="Plant_Dashboard", index=False,
                              startrow=len(dash_df) + 3 + len(diag_dist_df) + 3)
        summary_df.to_excel(writer, sheet_name="Summary", index=False)
        coverage_df.to_excel(writer, sheet_name="Data_Quality_Coverage", index=False)
        if not sel_df.empty:
            sel_df.to_excel(writer, sheet_name="Diagnostic_Selection", index=False)
        if show_stiction_sheet and not stiction_df.empty:
            stiction_df.to_excel(writer, sheet_name="Stiction_Analysis", index=False)
        if show_propagation_sheet and not prop_df.empty:
            prop_df.to_excel(writer, sheet_name="Propagation", index=False)
        if not dq_df.empty:
            dq_df.to_excel(writer, sheet_name="Per_Loop_Data_Quality", index=False)
        if not maint_df.empty:
            maint_df.to_excel(writer, sheet_name="Maintenance_Actions", index=False)
        cfg_doc_df.to_excel(writer, sheet_name="CONFIG_DOCUMENTATION", index=False)
        how_to.to_excel(writer, sheet_name="How_To_Read", index=False)

    # Style and embed dashboard image
    wb = load_workbook(path)
    for ws_name in wb.sheetnames:
        ws = wb[ws_name]
        try:
            _bold_header(ws)
            _autofit(ws)
        except Exception:
            pass
    if dashboard_png and os.path.exists(dashboard_png):
        ws = wb["Plant_Dashboard"]
        try:
            img = XLImage(dashboard_png)
            img.width = 800
            img.height = 280
            ws.add_image(img, "F2")
        except Exception:
            pass
    if heatmap_png and os.path.exists(heatmap_png):
        ws = wb["Plant_Dashboard"]
        try:
            img = XLImage(heatmap_png)
            # scale heatmap based on its aspect ratio
            img.width = 900
            img.height = 480
            ws.add_image(img, "F22")
        except Exception:
            pass
    wb.save(path)


_CONFIG_DOC = {
    "AMP_THRESHOLD": "PV peak-to-peak threshold for high oscillation flag.",
    "OP_ACTIVITY_THRESHOLD": "Mean |dOP| threshold for high OP activity flag.",
    "IAE_PER_HOUR_THRESHOLD": "IAE/hour threshold for poor tracking flag.",
    "STICT_CONF_HIGH": "Stiction consensus confidence (%) considered HIGH.",
    "STICT_CONF_MED": "Stiction consensus confidence (%) considered MEDIUM.",
    "PROP_CONF_MIN": "Min combined propagation score (%) to log a link.",
    "PROP_CONF_STRONG": "Combined propagation score (%) considered STRONG.",
    "CROSS_UNIT_DOWNWEIGHT": "Multiplier applied to combined propagation score for cross-unit pairs (per UNIT_MAPPING). 1.0 = no downweight; 0.5 = halve cross-unit scores.",
    "SERVICE_FACTOR_MIN_PCT": "Min % time in AUTO/CAS for loop to be analysed.",
    "SS_DETECTION_WINDOW": "Window size (samples) for steady-state detection.",
    "SS_STD_THRESHOLD": "Std threshold for steady-state classification.",
    "FROZEN_SAMPLES_MIN": "Min consecutive identical PV samples to flag frozen sensor.",
    "QUANTISATION_UNIQUE_VALS_MAX": "Max unique PV values to flag quantisation.",
    "COMPRESSION_FLAT_FRACTION_MAX": "Max fraction of compressed flat points before flagging.",
    "OSCILLATION_REGULARITY_MIN": "Min Hägglund regularity for oscillation flag.",
    "STICTION_S_MIN_PCT": "Min S (stickband %) to consider stiction physically present.",
    "HARRIS_INDEX_THRESHOLD": "Min Harris index for 'good control'.",
    "STIC_W_HEURISTIC": "Weight of heuristic method in stiction consensus.",
    "STIC_W_HORCH": "Weight of Horch CC method in stiction consensus.",
    "STIC_W_YAMASHITA": "Weight of Yamashita shape method in stiction consensus.",
    "STIC_W_BICOH": "Weight of bicoherence method in stiction consensus.",
    "MAX_DT_FOR_STICTION_SEC": "Max sample interval (sec) for stiction detection.",
    "MAX_DT_FOR_OSCILLATION_SEC": "Max sample interval (sec) for oscillation analysis.",
    "MAX_DT_FOR_HARRIS_SEC": "Max sample interval (sec) for Harris index.",
    "MAX_DT_FOR_PROPAGATION_SEC": "Max sample interval (sec) for cross-loop propagation.",
    "TOP_N_WORST_LOOPS": "How many worst-performing loops to list on dashboard.",
    "PLANT_HEALTH_GOOD_THRESHOLD": "Health score above which loop is 'Good'.",
    "PLANT_HEALTH_POOR_THRESHOLD": "Health score below which loop is 'Critical'.",
}
